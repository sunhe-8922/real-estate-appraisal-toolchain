#!/usr/bin/env python3
"""
rebuild_excel_formula.py — 从 JSON calculationChain 重建 Excel 公式（v1.2）

两种模式：
  --mode values  数值公式验证：将公式模板中的 {{refKey}} 替换为 JSON 数据实际值，
                 求值并与 target 字段值对比，报告 PASS/FAIL。
  --mode cells   单元格引用公式：将 {{refKey}} 替换回 Excel 单元格引用，
                 重建为 Excel 可用的公式字符串；可选 --excel 与原模板逐字对比。

用法：
  python scripts/rebuild_excel_formula.py --mode values --chain outputs/calculation_chain.json --data schema/example-武汉洪山住宅.json
  python scripts/rebuild_excel_formula.py --mode cells --chain outputs/calculation_chain.json [--excel outputs/房地产评估明细表-计算模板.xlsx]

cells 模式重建规则：
  {{refKey}}  → 按 refs[refKey] 的 JSONPath 查反向映射表，替换为单元格引用
  SUM({{refKey}}) → SUM(单元格区域)，如 SUM(V7:V18)
  数值常量（0.5/0.3/0.2）与布局常量（-1100/-400/-700）原样保留
  Excel 局部引用（如 G23）原样保留

values 模式求值规则：
  {{refKey}} → 按 JSONPath 从数据取值；数组（Details）→ 取 factor 数组并 SUM
  ROUND(x, n) → round(x, n)；^ → **
"""

import argparse
import ast
import json
import operator
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None


# ── 反向映射：JSONPath → Excel 单元格引用 ─────────────────────────
# 模板布局常量（住宅模板）。comparables 部分按实例索引 [0]/[1]/[2] 展开。
COMP_PARTS = [
    (0, "T", "V", 7, 31),
    (1, "W", "Y", 7, 31),
    (2, "Z", "AB", 7, 31),
]

INSTANCE_COLS = {0: ("T", "V"), 1: ("W", "Y"), 2: ("Z", "AB")}
INSTANCE_RANGES = {0: "V", 1: "Y", 2: "AB"}

# (JSONPath, 模板 sheet, 单元格引用模式)
REVERSE_MAP = [
    # 比较法实例
    ("methods.comps.comparableInstances[{i}].unitPrice", "市场价比较法", "{eval_col}4"),
    ("methods.comps.comparableInstances[{i}].adjustments.transactionSituation", "市场价比较法", "{eval_col}5/{inst_col}5"),
    ("methods.comps.comparableInstances[{i}].adjustments.marketCondition", "市场价比较法", "{eval_col}6/{inst_col}6"),
    ("methods.comps.comparableInstances[{i}].adjustments.locationDetails", "市场价比较法", "{inst_col}7:{inst_col}18"),
    ("methods.comps.comparableInstances[{i}].adjustments.interestDetails", "市场价比较法", "{inst_col}19:{inst_col}23"),
    ("methods.comps.comparableInstances[{i}].adjustments.physicalDetails", "市场价比较法", "{inst_col}24:{inst_col}31"),
    ("methods.comps.comparableInstances[{i}].adjustedUnitPrice", "市场价比较法", "{eval_col}32"),
    # 收益法
    ("methods.income.netOperatingIncome.annualAmount", "住宅-收益法测算", "G4"),
    ("methods.income.netOperatingIncome.effectiveGrossIncome", "住宅-收益法测算", "G5"),
    ("methods.income.netOperatingIncome.operatingExpenses", "住宅-收益法测算", "G11"),
    ("methods.income.netOperatingIncome.growthRate", "住宅-收益法测算", "G25"),
    ("methods.income.rate.value", "住宅-收益法测算", "G24"),
    ("methods.income.holdingPeriod", "住宅-收益法测算", "G26"),
    ("methods.income.finalValue.total", "住宅-收益法测算", "G27"),
    ("result.weightAllocation.income", "住宅-收益法测算", "I27"),
    ("result.weightAllocation.comps", "住宅-收益法测算", "I28"),
    ("methods.comps.finalValue.total", "住宅-收益法测算", "J28"),
    ("result.finalTotalValue", "住宅-收益法测算", "J29"),
    # 评估明细表
    ("property.area", "评估明细表", "M6"),
    ("result.finalUnitValue", "评估明细表", "N6"),
    ("result.finalTotalValue", "评估明细表", "O6"),
]


def _build_reverse_lookup() -> dict[str, str]:
    """构建 {JSONPath: 单元格引用} 查找表（实例部分展开）。"""
    lookup = {}
    for path, sheet, pattern in REVERSE_MAP:
        if "{i}" in path:
            for i, (eval_col, inst_col) in INSTANCE_COLS.items():
                p = path.replace("{i}", str(i))
                cell = f"{sheet}!{pattern.format(eval_col=eval_col, inst_col=inst_col)}"
                lookup[p] = cell
        else:
            lookup[path] = f"{sheet}!{pattern}"
    return lookup


REVERSE_LOOKUP = _build_reverse_lookup()

# 用于 cells 逐字校验：SUM(V7:V18) → V7+V8+...+V18
SUM_RANGE_RE = re.compile(r"SUM\(([A-Z]{1,2})(\d+):\1(\d+)\)")


def expand_sum(formula: str) -> str:
    """将 SUM(V7:V18) 展开为 V7+V8+...+V18（用于与 Excel 原公式逐字对比）。"""
    def repl(m):
        col, r1, r2 = m.group(1), int(m.group(2)), int(m.group(3))
        return "+".join(f"{col}{r}" for r in range(r1, r2 + 1))
    return SUM_RANGE_RE.sub(repl, formula)


def get_jsonpath(data: dict, path: str):
    """按点分 JSONPath 取值；数组索引 [n] 支持。返回 (ok, value)。"""
    parts = re.split(r"\.|\[(\d+)\]", path)
    cur = data
    for p in parts:
        if p is None or p == "":
            continue
        if isinstance(cur, dict) and p in cur:
            cur = cur[p]
        elif isinstance(cur, list) and p.isdigit() and int(p) < len(cur):
            cur = cur[int(p)]
        else:
            return False, None
    return True, cur


def resolve_ref_value(data: dict, path: str):
    """解析 refs 的值：数组（Details 对象列表）→ factor 列表；标量 → 原值。"""
    ok, val = get_jsonpath(data, path)
    if not ok:
        return None
    if isinstance(val, list):
        # Details 数组：取每项 factor（100 基准指数）
        factors = []
        for item in val:
            if isinstance(item, dict) and "factor" in item:
                factors.append(item["factor"])
        return factors
    return val


# ── 安全求值器（AST 白名单，替代裸 eval）──────────────────────────
# calculationChain 是外部 JSON（可由 AI/他人提供），必须按不可信输入处理。
# 白名单：数字/列表字面量、算术(+ - * / **)与比较(= <> < <= > >=)运算、
# 一元 ±、IF（惰性求值，只计算匹配分支）+
# round/sum/max/min/abs/average/power/int/sqrt/and/or。
# 其余一律拒绝（属性访问/下标/推导式/名称引用/lambda/布尔运算符 and/or 等）。
_SAFE_BINOPS = {
    ast.Add: operator.add,
    ast.Sub: operator.sub,
    ast.Mult: operator.mul,
    ast.Div: operator.truediv,
    ast.Pow: operator.pow,
}
_SAFE_UNARYOPS = {ast.UAdd: operator.pos, ast.USub: operator.neg}
_SAFE_CMPOPS = {
    ast.Eq: operator.eq,
    ast.NotEq: operator.ne,
    ast.Lt: operator.lt,
    ast.LtE: operator.le,
    ast.Gt: operator.gt,
    ast.GtE: operator.ge,
}
_SAFE_FUNCS = {
    "round": round, "ROUND": round,
    "sum": sum, "SUM": sum,
    "max": max, "MAX": max,
    "min": min, "MIN": min,
    "abs": abs, "ABS": abs,
    "AVERAGE": lambda *a: sum(a) / len(a) if a else 0,
    "POWER": pow,
    "INT": lambda x: int(x),
    "SQRT": lambda x: x ** 0.5,
    "AND": lambda *a: all(a),
    "OR": lambda *a: any(a),
}


def safe_eval(expr: str):
    """AST 白名单求值。非法结构（属性访问/下标/推导式/名称引用等）抛 ValueError。"""
    tree = ast.parse(expr, mode="eval")

    def walk(node):
        if isinstance(node, ast.Expression):
            return walk(node.body)
        if isinstance(node, ast.Constant):
            if isinstance(node.value, (int, float)):
                return node.value
            raise ValueError(f"非法常量: {node.value!r}")
        if isinstance(node, ast.List):
            return [walk(e) for e in node.elts]
        if isinstance(node, ast.BinOp) and type(node.op) in _SAFE_BINOPS:
            return _SAFE_BINOPS[type(node.op)](walk(node.left), walk(node.right))
        if isinstance(node, ast.UnaryOp) and type(node.op) in _SAFE_UNARYOPS:
            return _SAFE_UNARYOPS[type(node.op)](walk(node.operand))
        if isinstance(node, ast.Compare):
            if len(node.ops) != 1:
                raise ValueError("不支持链式比较")
            op_type = type(node.ops[0])
            if op_type not in _SAFE_CMPOPS:
                raise ValueError(f"非法比较运算符: {op_type.__name__}")
            return _SAFE_CMPOPS[op_type](walk(node.left), walk(node.comparators[0]))
        if isinstance(node, ast.Call):
            if not isinstance(node.func, ast.Name) or node.keywords:
                raise ValueError("仅允许无关键字参数的白名单函数调用")
            fname = node.func.id
            # IF 惰性求值：只计算匹配分支（Excel 语义，避免未取分支的副作用）
            if fname in ("IF", "if"):
                if len(node.args) < 2:
                    raise ValueError("IF 需要 ≥2 个参数")
                cond = walk(node.args[0])
                if cond:
                    return walk(node.args[1])
                if len(node.args) >= 3:
                    return walk(node.args[2])
                return False
            if fname not in _SAFE_FUNCS:
                raise ValueError(f"函数不在白名单: {fname}")
            return _SAFE_FUNCS[fname](*[walk(a) for a in node.args])
        raise ValueError(f"非法语法节点: {type(node).__name__}")

    return walk(tree)


def rebuild_values(chain: dict, data: dict) -> list[dict]:
    """values 模式：数值公式求值验证。返回每个节点的验证结果。"""
    # 节点级容差：默认 ±1；ROUND(...,-1) 到十位 → ±10；
    # result.totalValue 为 面积×单价 vs 加权总价 的双口径舍入传播 → ±65。
    NODE_TOLERANCE = {
        "comps.finalUnitPrice": 10,
        "result.finalTotalValue": 10,
        "result.totalValue": 65,
    }
    results = []

    def eval_expr(expr: str, refs: dict) -> float:
        # {{refKey}} → 数据值
        def repl(m):
            key = m.group(1)
            path = refs.get(key)
            if not path:
                raise ValueError(f"refs 中无 {key}")
            val = resolve_ref_value(data, path)
            if val is None:
                raise ValueError(f"数据中无 {path}")
            if isinstance(val, list):
                return "[" + ",".join(str(v) for v in val) + "]"
            return str(val)
        body = re.sub(r"\{\{(\w+)\}\}", repl, expr)
        body = body.lstrip("=")  # 公式模板可能带前导 =
        # 语言转换：^→**、<>→!=（Excel 语法 → Python）
        body = body.replace("^", "**")
        body = body.replace("<>", "!=")
        # 安全求值：AST 白名单，拒绝一切非数值结构（防御恶意 formula/注入值）
        return safe_eval(body)

    for node in chain["nodes"]:
        formula = node["formula"]
        refs = node.get("refs", {})
        target_path = node["target"]
        try:
            computed = eval_expr(formula, refs)
        except Exception as e:  # noqa: BLE001
            results.append({
                "id": node["id"], "status": "SKIP",
                "reason": f"求值失败: {e}", "target": target_path,
            })
            continue
        ok, actual = get_jsonpath(data, target_path)
        if not ok or actual is None:
            results.append({
                "id": node["id"], "status": "SKIP",
                "reason": f"数据无 target 字段 {target_path}", "computed": computed,
            })
            continue
        # 容差：ROUND 舍入误差 + 节点级口径差异
        tolerance = NODE_TOLERANCE.get(node["id"], 1)
        diff = abs(round(float(computed)) - float(actual))
        passed = diff <= tolerance
        results.append({
            "id": node["id"], "status": "PASS" if passed else "FAIL",
            "computed": round(float(computed), 4), "actual": actual, "diff": diff,
            "tolerance": tolerance,
        })
    return results


def rebuild_cells(chain: dict) -> list[dict]:
    """cells 模式：单元格引用公式重建。返回每个节点的重建结果。"""
    results = []
    for node in chain["nodes"]:
        formula = node["formula"]
        refs = node.get("refs", {})
        # {{refKey}} → 单元格（在 SUM(...) 内部则替换为区域引用）
        def repl(m):
            key = m.group(1)
            path = refs.get(key)
            if not path:
                return m.group(0)
            cell = REVERSE_LOOKUP.get(path)
            if not cell:
                return m.group(0)
            return cell.split("!", 1)[1]
        rebuilt = re.sub(r"\{\{(\w+)\}\}", repl, formula)
        results.append({
            "id": node["id"], "excelSource": node["excelSource"],
            "rebuilt": rebuilt,
        })
    return results


def verify_against_excel(chain: dict, excel_path: str) -> list[dict]:
    """cells 模式校验：重建公式（展开 SUM）与 Excel 原公式归一化后逐字对比。

    归一化规则：Excel 原公式中未被重建公式引用的单元格，
    若为数值常量 → 替换为字面量（如 T33 权重 0.5）；
    若为简单中转引用（=+X 或 =X）→ 替换为目标单元格（如 J27→G27）。
    """
    if openpyxl is None:
        return [{"error": "openpyxl 未安装，无法校验"}]
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    checks = []
    for node in chain["nodes"]:
        sheet, cell = node["excelSource"].split("!")
        ws = wb[sheet]
        original = ws[cell].value
        rebuilt_cells = rebuild_cells({"nodes": [node]})[0]["rebuilt"]
        expanded = expand_sum(rebuilt_cells)

        # 重建公式中引用的单元格集合（本 sheet 无前缀）
        rebuilt_cell_refs = set(re.findall(r"(?<![A-Z])([A-Z]{1,2}\d+)", expanded))
        # 归一化 Excel 原公式
        normalized = normalize_excel_formula(
            str(original).lstrip("="), sheet, wb, rebuilt_cell_refs, depth=0
        )
        match = normalized == expanded.lstrip("=")
        checks.append({
            "id": node["id"], "match": match,
            "excelSource": node["excelSource"],
            "original": str(original),
            "normalized": normalized,
            "rebuilt": expanded.lstrip("="),
        })
    wb.close()
    return checks


def normalize_excel_formula(formula: str, sheet: str, wb, keep_refs: set, depth: int = 0) -> str:
    """将 Excel 公式中未被重建引用的单元格展开（常量→字面量、中转→目标）。"""
    if depth > 5:
        return formula

    def repl(m):
        cell = m.group(0).replace("$", "")
        # 带 sheet 前缀的引用（如 市场价比较法!T32）保持原样——重建公式用本 sheet 名
        if "!" in cell:
            return m.group(0)
        ws = wb[sheet]
        val = ws[cell].value
        if val is None or cell in keep_refs:
            return m.group(0)
        if isinstance(val, (int, float)):
            # 数值常量 → 字面量
            if isinstance(val, float) and val.is_integer():
                return str(int(val))
            return str(val)
        if isinstance(val, str) and val.startswith("="):
            # 简单中转 =+X 或 ==+X → 目标单元格（递归展开）
            target = val.lstrip("=").lstrip("+").strip()
            if re.fullmatch(r"([A-Z]{1,2}\d+)", target):
                return normalize_excel_formula(target, sheet, wb, keep_refs, depth + 1)
        return m.group(0)

    return re.sub(r"(?<![A-Z])(\$?[A-Z]{1,2}\$?\d+)", repl, formula)


def main():
    parser = argparse.ArgumentParser(description="JSON calculationChain → Excel 公式重建")
    parser.add_argument("--mode", choices=["values", "cells"], required=True)
    parser.add_argument("--chain", default="outputs/calculation_chain.json")
    parser.add_argument("--data", default="schema/example-武汉洪山住宅.json",
                        help="values 模式的数据 JSON")
    parser.add_argument("--excel", default="outputs/房地产评估明细表-计算模板.xlsx",
                        help="cells 模式校验用 Excel 模板")
    args = parser.parse_args()

    chain_path = Path(args.chain)
    if not chain_path.exists():
        print(f"ERROR: 计算链不存在: {chain_path}")
        sys.exit(1)
    with open(chain_path, encoding="utf-8") as f:
        chain = json.load(f)

    if args.mode == "values":
        data_path = Path(args.data)
        if not data_path.exists():
            print(f"ERROR: 数据不存在: {data_path}")
            sys.exit(1)
        with open(data_path, encoding="utf-8") as f:
            data = json.load(f)
        print(f"values 模式：{chain_path.name} + {data_path.name}")
        print(f"{'节点':<36}{'状态':<6}{'计算值':<14}{'实际值':<14}差异")
        print("-" * 90)
        passed = 0
        for r in rebuild_values(chain, data):
            if r["status"] == "PASS":
                passed += 1
            diff_col = r.get('diff', r.get('reason', ''))
            tol_col = f"(±{r['tolerance']})" if 'tolerance' in r else ""
            print(f"{r['id']:<36}{r['status']:<6}"
                  f"{str(r.get('computed', '')):<14}{str(r.get('actual', '')):<14}"
                  f"{diff_col} {tol_col}")
        total = len(chain["nodes"])
        print("-" * 90)
        print(f"PASS {passed}/{total}")
        sys.exit(0 if passed == total else 1)

    # cells 模式
    print(f"cells 模式：{chain_path.name}")
    print(f"{'节点':<36}{'excelSource':<22}重建公式")
    print("-" * 110)
    for r in rebuild_cells(chain):
        print(f"{r['id']:<36}{r['excelSource']:<22}{r['rebuilt']}")
    print()

    if args.excel and Path(args.excel).exists():
        print(f"逐字校验（vs {Path(args.excel).name}，归一化后对比）：")
        checks = verify_against_excel(chain, args.excel)
        ok = all(c.get("match", False) for c in checks)
        for c in checks:
            if "error" in c:
                print(f"  ERROR: {c['error']}")
                continue
            mark = "✓" if c["match"] else "✗"
            print(f"  {mark} {c['id']} ({c['excelSource']})")
            if not c["match"]:
                print(f"      Excel: {c['original']}")
                print(f"      归一化: {c['normalized']}")
                print(f"      重建  : {c['rebuilt']}")
        print(f"\n逐字校验：{sum(1 for c in checks if c.get('match'))}/{len(checks)} 一致")
    else:
        print("（未提供 --excel，跳过逐字校验）")


if __name__ == "__main__":
    main()
