"""
Excel 模板 → JSON Schema 字段映射验证脚本

从 Excel 计算模板提取字段，构建 appraisal-result 格式的 JSON 对象，
验证映射可行性和识别缺口。

用法：
    python scripts/excel_to_json_mapping.py [--excel PATH] [--output PATH]
"""

import json
import sys
import argparse
from pathlib import Path
from datetime import datetime, date

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


def excel_serial_to_date(serial):
    """Excel 序列号 → YYYY-MM-DD"""
    if isinstance(serial, (int, float)):
        base = datetime(1899, 12, 30)
        return (base + timedelta(days=serial)).strftime("%Y-%m-%d")
    elif isinstance(serial, datetime):
        return serial.strftime("%Y-%m-%d")
    return str(serial)


def parse_prefix(value, prefix):
    """解析带前缀的单元格值，如 '委托人：XX公司' → 'XX公司'"""
    if value and isinstance(value, str) and prefix in value:
        return value.split(prefix)[-1].strip()
    return value


from datetime import timedelta


def extract_from_excel(excel_path):
    """从 Excel 模板提取字段，返回 (mapped_dict, gaps_list)"""

    wb = openpyxl.load_workbook(excel_path, data_only=False)
    mapped = {}
    gaps = []

    # ── project ──
    ws = wb["评估明细表"]
    mapped["project"] = {
        "client": parse_prefix(ws["B3"].value, "委托人："),
        "agency": parse_prefix(ws["B9"].value, "评估机构："),
        "appraiser": {
            "name": parse_prefix(ws["J9"].value, "评估人员："),
        },
    }
    # reportDate 需解析中文日期
    raw_date = ws["B10"].value
    if raw_date and isinstance(raw_date, str):
        # "日期：2028年08月11日" → "2028-08-11"
        import re
        m = re.search(r"(\d{4})年(\d{2})月(\d{2})日", raw_date)
        if m:
            mapped["project"]["reportDate"] = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"

    for field, label in [("purpose", "估价目的"), ("name", "项目名称"), ("reportNumber", "报告编号")]:
        gaps.append(f"project.{field} — Excel 无 '{label}' 字段")
    gaps.append("project.appraiser.registrationNumber — Excel 无 '注册号' 字段")

    # ── property ──
    ws = wb["评估明细表"]
    completion_dt = ws["K6"].value
    if isinstance(completion_dt, datetime):
        completion_year = completion_dt.year
    elif isinstance(completion_dt, (int, float)):
        completion_year = excel_serial_to_date(completion_dt)[:4]
    else:
        completion_year = None

    floor_str = ""
    if ws["I6"].value and ws["J6"].value:
        floor_str = f"{ws['J6'].value}/{ws['I6'].value}"

    mapped["property"] = {
        "area": ws["M6"].value,
        "usage": ws["F6"].value,
        "address": ws["D6"].value,
        "completionYear": completion_year,
        "floor": floor_str,
        "decoration": ws["G6"].value,
    }

    # orientation from 住宅-实物状况
    ws_phys = wb["住宅-实物状况"]
    mapped["property"]["orientation"] = ws_phys["L4"].value

    # remainingUsefulLife / landUseRightYears — 需读取收益法公式结果
    ws_income = wb["住宅-收益法测算"]
    # 公式值无法直接读取（data_only=False），记录公式
    j11_formula = ws_income["J11"].value
    j6_formula = ws_income["J6"].value
    mapped["property"]["_remainingUsefulLife_formula"] = str(j11_formula) if j11_formula else None
    mapped["property"]["_landUseRightYears_formula"] = str(j6_formula) if j6_formula else None
    gaps.append("property.remainingUsefulLife — Excel 为公式，需 data_only=True 读取计算值")
    gaps.append("property.landUseRightYears — Excel 为公式，需 data_only=True 读取计算值")
    gaps.append("property.propertyType — Excel 无 '房地产类型' 字段")
    gaps.append("property.ownershipType — Excel 无 '权属类型' 字段")

    # ── valuation ──
    ws = wb["评估明细表"]
    value_date_serial = ws["J2"].value
    mapped["valuation"] = {
        "valueDate": excel_serial_to_date(value_date_serial) if value_date_serial else None,
        "currency": "CNY",
    }
    gaps.append("valuation.valueType — Excel 无 '价值类型' 字段")

    # ── methods.comps ──
    ws_cmp = wb["市场价比较法"]
    instances = []
    for col_idx, col_letter in enumerate(["E", "F", "G"], 1):
        name = ws_cmp[f"{col_letter}1"].value
        location = ws_cmp[f"{col_letter}2"].value
        unit_price = ws_cmp[f"{col_letter}4"].value
        area = ws_cmp[f"{col_letter}24"].value
        tx_date_serial = ws_cmp[f"{col_letter}6"].value

        # 修正系数（100 基准 → 小数）
        tx_situation_raw = ws_cmp[f"{col_letter}5"].value  # 交易情况原始值
        # 修正系数在 M/N/O 列（L=评估对象, M=实例1, N=实例2, O=实例3）
        adj_cols = ["M", "N", "O"]
        adj_col = adj_cols[col_idx - 1]

        # 修正系数（100 基准指数 → 小数倍率 = 评估对象指数/实例指数，忠实 Excel T32: T5/V5）
        tx_situation_coef = ws_cmp[f"{adj_col}5"].value
        market_cond_coef = ws_cmp[f"{adj_col}6"].value
        eval_obj_tx = ws_cmp[f"L5"].value or 100
        eval_obj_mkt = ws_cmp[f"L6"].value or 100

        # 区位状况子项：行 7-18 (12项) — 对应 Excel 公式 100/(V7+...+V18-1100)
        location_factors = []
        location_details = []
        for row in range(7, 19):
            v = ws_cmp[f"{adj_col}{row}"].value
            label = ws_cmp[f"D{row}"].value
            if v is not None and isinstance(v, (int, float)):
                location_factors.append(v)
                location_details.append({"name": str(label) if label else f"区位{row-6}", "factor": v})

        # 权益状况子项：行 19-23 (5项) — 对应 Excel 公式 100/(V19+...+V23-400)
        interest_factors = []
        interest_details = []
        for row in range(19, 24):
            v = ws_cmp[f"{adj_col}{row}"].value
            label = ws_cmp[f"D{row}"].value
            if v is not None and isinstance(v, (int, float)):
                interest_factors.append(v)
                interest_details.append({"name": str(label) if label else f"权益{row-18}", "factor": v})

        # 实物状况子项：行 24-31 (8项) — 对应 Excel 公式 100/(V24+...+V31-700)
        physical_factors = []
        physical_details = []
        for row in range(24, 32):
            v = ws_cmp[f"{adj_col}{row}"].value
            label = ws_cmp[f"D{row}"].value
            if v is not None and isinstance(v, (int, float)):
                physical_factors.append(v)
                physical_details.append({"name": str(label) if label else f"实物{row-23}", "factor": v})

        # 合并子项：Excel 偏差加总法 100/(100+Σ(factor-100))，忠实 市场价比较法!T32 公式
        # （注意：不是连乘！连乘会得到不同数值）
        def merge_factors(factors):
            if not factors:
                return None
            deviation = sum(f - 100 for f in factors)
            return round(100 / (100 + deviation), 6)

        # 比准单价
        adjusted_price_cols = ["T", "W", "Z"]
        adj_col_price = adjusted_price_cols[col_idx - 1]
        adjusted_unit_price_formula = ws_cmp[f"{adj_col_price}32"].value

        instance = {
            "name": name,
            "location": location,
            "area": area,
            "unitPrice": unit_price,
            "transactionPrice": round(unit_price * area, 2) if unit_price and area else None,
            "transactionDate": excel_serial_to_date(tx_date_serial) if tx_date_serial else None,
            "adjustments": {
                "transactionSituation": round(eval_obj_tx / tx_situation_coef, 6) if tx_situation_coef else None,
                "marketCondition": round(eval_obj_mkt / market_cond_coef, 6) if market_cond_coef else None,
                "location": merge_factors(location_factors),
                "physical": merge_factors(physical_factors),
                "interest": merge_factors(interest_factors),
                "locationDetails": location_details,
                "physicalDetails": physical_details,
                "interestDetails": interest_details,
            },
            "adjustedUnitPrice_formula": str(adjusted_unit_price_formula) if adjusted_unit_price_formula else None,
        }
        instances.append(instance)

    # 权重
    weight_cols = ["T", "W", "Z"]
    weights = []
    for col in weight_cols:
        w = ws_cmp[f"{col}33"].value
        if w is not None:
            weights.append(w)

    # 最终单价
    final_unit_formula = ws_cmp["T34"].value

    mapped["methods"] = {
        "comps": {
            "applicable": True,
            "comparableInstances": instances,
            "finalValue": {
                "unit_formula": str(final_unit_formula) if final_unit_formula else None,
                "total_formula": "=T34 * 评估明细表!M6" if final_unit_formula else None,
            },
            "weight": weights,
            "_weight_note": "权重在 T33/W33/Z33，方法间权重在 住宅-收益法!I27/I28",
        },
    }
    gaps.append("comps.weightRationale — Excel 无权重理由文字")
    gaps.append("comps.redLineChecks[] — Excel 无红线检查记录")
    gaps.append("comps.comparableInstances[].sourceGrade — Excel 无信源等级标注")
    gaps.append("comps.comparableInstances[].adjustedUnitPrice — Excel 为公式，需 data_only=True")

    # ── methods.income ──
    ws_inc = wb["住宅-收益法测算"]
    ws_rate = wb["住宅-收益率"]

    mapped["methods"]["income"] = {
        "applicable": True,
        "calculationMode": "fullRemainingLife",  # 从公式结构推断
        "incomeType": "rentalIncome",
        "netOperatingIncome": {
            "effectiveGrossIncome_formula": str(ws_inc["G5"].value) if ws_inc["G5"].value else None,
            "operatingExpenses_formula": str(ws_inc["G11"].value) if ws_inc["G11"].value else None,
            "annualAmount_formula": str(ws_inc["G4"].value) if ws_inc["G4"].value else None,
            "growthRate": ws_inc["G25"].value,
        },
        "rate": {
            "type": "yieldRate",
            "value_formula": str(ws_inc["G24"].value) if ws_inc["G24"].value else None,
            "determinationMethod": "累加法（安全利率+风险补偿+管理负担补偿）",
        },
        "finalValue": {
            "unit_formula": str(ws_inc["G27"].value) if ws_inc["G27"].value else None,
        },
    }
    gaps.append("income.netOperatingIncome.* — Excel 为公式，需 data_only=True 读取计算值")
    gaps.append("income.rate.value — Excel 为公式（='住宅-收益率'!O14），需 data_only=True")
    gaps.append("income.weightRationale — Excel 无权重理由")
    gaps.append("income.redLineChecks[] — Excel 无红线检查")
    gaps.append("income.netOperatingIncome.historicalDataYears — Excel 无历史数据年数")

    # ── methods.cost / hypotheticalDev ──
    mapped["methods"]["cost"] = None
    mapped["methods"]["hypotheticalDev"] = None
    gaps.append("methods.cost — Excel 住宅模板不含成本法")
    gaps.append("methods.hypotheticalDev — Excel 住宅模板不含假设开发法")

    # ── result ──
    ws = wb["评估明细表"]
    mapped["result"] = {
        "finalUnitValue": ws["N6"].value,
        "finalTotalValue_formula": str(ws["O6"].value) if ws["O6"].value else None,
    }
    gaps.append("result.determinationMethod — Excel 无结果确定方式描述")
    gaps.append("result.weightAllocation — 需从收益法 sheet 推断")
    gaps.append("result.finalTotalValueInWords — Excel 无大写金额")
    gaps.append("result.crossMethodDifference — Excel 无差异分析")
    gaps.append("result.calculationMode — Excel 无计算模式 enum")

    # ── crossMethodConsistency ──
    mapped["crossMethodConsistency"] = []
    gaps.append("crossMethodConsistency[] — Excel 无跨方法一致性检查")

    # ── crossMethodNotes ──
    gaps.append("crossMethodNotes — Excel 无跨方法讨论笔记")

    wb.close()
    return mapped, gaps


def main():
    parser = argparse.ArgumentParser(description="Excel 模板 → JSON Schema 字段映射验证")
    parser.add_argument("--excel", default="outputs/房地产评估明细表-计算模板.xlsx",
                        help="Excel 模板路径")
    parser.add_argument("--output", default="outputs/excel_to_json_mapping_preview.json",
                        help="输出 JSON 预览文件路径")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        print(f"ERROR: Excel 文件不存在: {excel_path}")
        sys.exit(1)

    print(f"读取 Excel: {excel_path}")
    mapped, gaps = extract_from_excel(excel_path)

    # 输出映射结果
    output_path = Path(args.output)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(mapped, f, ensure_ascii=False, indent=2)
    print(f"\n映射结果已保存: {output_path}")

    # 统计
    def count_fields(obj):
        count = 0
        if isinstance(obj, dict):
            for k, v in obj.items():
                if k.startswith("_"):
                    continue
                if v is not None:
                    count += 1
                count += count_fields(v)
        elif isinstance(obj, list):
            for item in obj:
                count += count_fields(item)
        return count

    mapped_count = count_fields(mapped)
    gap_count = len(gaps)

    print(f"\n{'='*60}")
    print(f"映射统计")
    print(f"{'='*60}")
    print(f"已映射字段数: {mapped_count}")
    print(f"缺口字段数:   {gap_count}")
    print(f"缺口比例:     {gap_count/(mapped_count+gap_count)*100:.1f}%")

    print(f"\n{'='*60}")
    print(f"缺口清单 ({gap_count} 项)")
    print(f"{'='*60}")
    for i, gap in enumerate(gaps, 1):
        print(f"  {i:2d}. {gap}")

    print(f"\n{'='*60}")
    print(f"关键发现")
    print(f"{'='*60}")
    print("""
    1. 核心数值字段（面积/单价/总价/可比实例数据）可直接映射
    2. 日期需转换（Excel serial → ISO date）
    3. 修正系数需转换（100 基准 → 小数）
    4. 区位/实物/权益修正需合并多子项为单一系数
    5. 公式字段需 data_only=True 二次读取计算值
    6. 18 个 Schema 必填/可选字段在 Excel 中无对应
    7. 成本法/假设开发法在住宅模板中无对应 sheet
    """)


if __name__ == "__main__":
    main()
