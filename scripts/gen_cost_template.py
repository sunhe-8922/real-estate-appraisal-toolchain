#!/usr/bin/env python3
"""生成成本法测算 Excel 模板 — GB/T 50291-2015 第4.4节"""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from xlsx_meta import save_frozen
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

# ── 颜色定义 ──
BLUE_FONT = Font(name="Arial", size=11, color="0000FF")
BLACK_FONT = Font(name="Arial", size=11, color="000000")
BOLD_BLACK = Font(name="Arial", size=11, color="000000", bold=True)
TITLE_FONT = Font(name="Arial", size=14, color="FFFFFF", bold=True)
LEGEND_FONT = Font(name="Arial", size=9, color="333333", italic=True)

DARK_BLUE_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "outputs", "templates")
os.makedirs(OUTPUT_DIR, exist_ok=True)

wb = Workbook()

# ═══════════════════════════════════════════════
# SHEET 1: 重置成本
# ═══════════════════════════════════════════════
ws1 = wb.active
ws1.title = "重置成本"
ws1.column_dimensions['A'].width = 8
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 22
ws1.column_dimensions['D'].width = 35

# 标题
ws1.merge_cells('A1:D1')
c = ws1['A1']; c.value = "重置成本/重建成本测算"; c.font = TITLE_FONT; c.fill = DARK_BLUE_FILL; c.alignment = CENTER
ws1.row_dimensions[1].height = 30

# 项目信息
ws1.merge_cells('A2:B2')
ws1['A2'] = "估价对象: [请输入地址]"; ws1['A2'].font = BLUE_FONT
ws1.merge_cells('C2:D2')
ws1['C2'] = "价值时点: [请输入日期]     路径: 房地合估     成本类型: 重置成本"
ws1['C2'].font = Font(name="Arial", size=10, color="666666")

# 颜色图例
ws1.merge_cells('A3:D3')
ws1['A3'] = "图例: 蓝色字=可修改输入  |  黑色字=公式自动计算  |  黄色底=关键参数"
ws1['A3'].font = LEGEND_FONT; ws1['A3'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# 表头
for i, h in enumerate(["序号", "构成项", "金额 (元/m²)", "测算依据"], 1):
    cell = ws1.cell(row=5, column=i, value=h)
    cell.font = Font(name="Arial", size=11, color="000000", bold=True)
    cell.fill = LIGHT_BLUE_FILL; cell.alignment = CENTER; cell.border = THIN_BORDER

# 数据行
s1_data = [
    (6,  1, "土地成本",            "BLUE",    None,     "方法: [比较法/成本法/基准地价法] 4.4.6条"),
    (7,  2, "建设成本",            "BLUE",    None,     "方法: [单位比较法/分部分项法/工料测量法] 4.4.6条"),
    (8,  3, "管理费用",            "FORMULA", '=ROUND(C7*$F$8,0)', "公式: 管理费率($F$8) × 建设成本"),
    (9,  4, "销售费用",            "FORMULA", '=ROUND(C14*$F$9,0)', "公式: 售价参考行C14 × 销售费率($F$9)"),
    (10, 5, "投资利息",            "FORMULA", '=ROUND((C6+C7)*$F$10,0)', "公式: (土地成本+建设成本) × 年利率($F$10). 不含C9/C11. 4.4.6条第2款"),
    (11, 6, "销售税费",            "FORMULA", '=ROUND(C14*$F$11,0)', "公式: 售价参考行C14 × 税率($F$11)"),
    (12, 7, "开发利润",            "BLUE_Y",  None,     "计算基数×利润率($F$12). 4.4.6条. 黄底=关键参数"),
]

for row, seq, name, style, default, basis in s1_data:
    ws1.cell(row=row, column=1, value=seq).font = BLACK_FONT
    ws1.cell(row=row, column=1).alignment = CENTER; ws1.cell(row=row, column=1).border = THIN_BORDER

    ws1.cell(row=row, column=2, value=name).font = BOLD_BLACK
    ws1.cell(row=row, column=2).border = THIN_BORDER

    c3 = ws1.cell(row=row, column=3)
    if style == "BLUE":
        c3.font = BLUE_FONT; c3.value = None; c3.number_format = '#,##0'
    elif style == "BLUE_Y":
        c3.font = BLUE_FONT; c3.value = None; c3.number_format = '#,##0'; c3.fill = YELLOW_FILL
    elif style == "FORMULA":
        c3.font = BLACK_FONT; c3.value = default; c3.number_format = '#,##0'
    c3.border = THIN_BORDER; c3.alignment = CENTER

    ws1.cell(row=row, column=4, value=basis).font = Font(name="Arial", size=9, color="666666", italic=True)
    ws1.cell(row=row, column=4).alignment = LEFT_WRAP; ws1.cell(row=row, column=4).border = THIN_BORDER

# 附加参数行 (管理费率/销售费率/利率/税率/利润率) — F列存放实际参数值供公式引用
param_data = [
    (8,  "管理费率", 0.03, "0.0%"),
    (9,  "销售费率", 0.02, "0.0%"),
    (10, "利率",     0.04, "0.00%"),
    (11, "税率",     0.05, "0.0%"),
    (12, "利润率",   0.12, "0.0%"),
]
# 在E列放参数标签，F列放参数值
ws1.column_dimensions['E'].width = 14
ws1.column_dimensions['F'].width = 12
for row, pname, pdefault, fmt in param_data:
    ws1.cell(row=row, column=5, value=pname).font = Font(name="Arial", size=9, color="FF0000")
    ws1.cell(row=row, column=5).alignment = CENTER
    c = ws1.cell(row=row, column=6)
    c.font = BLUE_FONT; c.value = pdefault; c.number_format = fmt

# 合计行
row_total = 14
ws1.merge_cells(f'A{row_total}:B{row_total}')
ws1.cell(row=row_total, column=1, value="重置成本合计 (元/m²)").font = Font(name="Arial", size=12, color="000000", bold=True)
ws1.cell(row=row_total, column=1).alignment = CENTER
ws1.cell(row=row_total, column=1).border = THIN_BORDER

ws1.cell(row=row_total, column=3, value='=ROUND(SUM(C6:C12),0)').font = Font(name="Arial", size=12, color="000000", bold=True)
ws1.cell(row=row_total, column=3).number_format = '#,##0'; ws1.cell(row=row_total, column=3).fill = YELLOW_FILL
ws1.cell(row=row_total, column=3).alignment = CENTER

# 注释
comment_map = {
    (6, 3): "4.4.6条第1款: 土地成本 = 土地取得成本 + 土地开发成本。优先用比较法评估。",
    (7, 3): "4.4.6条第2款: 建设成本可按单位比较法、分部分项法或工料测量法求取。",
    (8, 3): "管理费用 = 建设成本 × 管理费率。管理费率参照同类项目或行业标准。管理费率在E8:F8输入。",
    (10, 3): "4.4.6条第2款: 投资利息计息基数不含销售费用和销售税费。各项计息期分别计算。利率在E10:F10输入。",
    (12, 3): "开发利润计算基数通常为土地成本+建设成本+管理费用+投资利息+销售费用。利润率在E12:F12输入。",
}
for pos, text in comment_map.items():
    ws1.cell(row=pos[0], column=pos[1]).comment = Comment(text, "GB/T 50291-2015")

# 数据验证
dv_positive = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0")
dv_positive.error = "数值必须≥0"; ws1.add_data_validation(dv_positive)
for row in [6, 7, 12]:
    dv_positive.add(ws1.cell(row=row, column=3))
for row in [8, 9, 10, 11, 12]:
    dv_positive.add(ws1.cell(row=row, column=6))

# ═══════════════════════════════════════════════
# SHEET 2: 折旧测算
# ═══════════════════════════════════════════════
ws2 = wb.create_sheet("折旧测算")
ws2.column_dimensions['A'].width = 28
ws2.column_dimensions['B'].width = 22
ws2.column_dimensions['C'].width = 28

# 标题
ws2.merge_cells('A1:C1')
c = ws2['A1']; c.value = "建筑物折旧测算"; c.font = TITLE_FONT; c.fill = DARK_BLUE_FILL; c.alignment = CENTER
ws2.row_dimensions[1].height = 30

# 参数区
ws2.merge_cells('A2:C2')
ws2['A2'] = "折旧方法选择 →"; ws2['A2'].font = BOLD_BLACK

for i, h in enumerate(["参数", "数值", "说明"], 1):
    cell = ws2.cell(row=3, column=i, value=h)
    cell.font = Font(name="Arial", size=11, color="000000", bold=True)
    cell.fill = LIGHT_BLUE_FILL; cell.alignment = CENTER; cell.border = THIN_BORDER

# 折旧方法下拉框
ws2.cell(row=4, column=1, value="折旧方法").font = BOLD_BLACK; ws2.cell(row=4, column=1).border = THIN_BORDER
ws2.cell(row=4, column=2).fill = YELLOW_FILL; ws2.cell(row=4, column=2).border = THIN_BORDER
ws2.cell(row=4, column=3, value="年龄-寿命法 / 成新折扣法 / 分解法").font = Font(name="Arial", size=9, color="666666")
ws2.cell(row=4, column=3).border = THIN_BORDER

dv_method = DataValidation(type="list", formula1='"年龄-寿命法,成新折扣法,分解法"', allow_blank=False)
dv_method.error = "请选择折旧方法"; dv_method.errorTitle = "方法选择"
ws2.add_data_validation(dv_method); dv_method.add(ws2['B4'])

# 参数行
param_rows = [
    (5, "重置成本 C (元/m²)",  "=重置成本!C7",     "仅建筑物部分; 跨表引用", False),
    (6, "预计净残值 S (元/m²)", "BLUE",             "建筑物拆除后可回收的净残值", False),
    (7, "有效年龄 t (年)",     "BLUE_Y",           "现场查勘判断(4.4.14条), 不等于建成年代", True),
    (8, "经济寿命 N (年)",     "BLUE",             "土地使用权届满则取剩余年限", False),
    (9, "成新率 q",             "BLUE",             "仅成新折扣法需填; 范围0~1", False),
]
for row, label, val, note, is_yellow in param_rows:
    ws2.cell(row=row, column=1, value=label).font = BOLD_BLACK; ws2.cell(row=row, column=1).border = THIN_BORDER
    c2 = ws2.cell(row=row, column=2)
    if val == "BLUE":
        c2.font = BLUE_FONT; c2.number_format = '#,##0'; c2.border = THIN_BORDER
    elif val == "BLUE_Y":
        c2.font = BLUE_FONT; c2.number_format = '#,##0'; c2.fill = YELLOW_FILL; c2.border = THIN_BORDER
        c2.comment = Comment("4.4.14条: 必须到现场查勘判断实际新旧程度，不得仅凭建成年代推算。", "GB/T 50291-2015")
    else:
        c2.font = BLACK_FONT; c2.value = val; c2.number_format = '#,##0'; c2.border = THIN_BORDER
    ws2.cell(row=row, column=3, value=note).font = Font(name="Arial", size=9, color="666666")
    ws2.cell(row=row, column=3).border = THIN_BORDER; ws2.cell(row=row, column=3).alignment = LEFT_WRAP

# 折旧结果区
row_sep = 11
ws2.merge_cells(f'A{row_sep}:C{row_sep}')
ws2.cell(row=row_sep, column=1, value="折旧结果").font = BOLD_BLACK

for i, h in enumerate(["折旧类型", "金额 (元/m²)", "原因"], 1):
    cell = ws2.cell(row=row_sep+1, column=i, value=h)
    cell.font = Font(name="Arial", size=11, color="000000", bold=True)
    cell.fill = LIGHT_BLUE_FILL; cell.alignment = CENTER; cell.border = THIN_BORDER

# 年龄-寿命法
ws2.cell(row=row_sep+2, column=1, value="直线法折旧(年龄-寿命)").font = BOLD_BLACK; ws2.cell(row=row_sep+2, column=1).border = THIN_BORDER
ws2.cell(row=row_sep+2, column=2, value="=ROUND((B5-B6)*MAX(0,MIN(1,B7/B8)),0)").font = BLACK_FONT
ws2.cell(row=row_sep+2, column=2).number_format = '#,##0'; ws2.cell(row=row_sep+2, column=2).border = THIN_BORDER
ws2.cell(row=row_sep+2, column=3, value="D = (C-S) × t/N。有效年龄≤经济寿命").font = Font(name="Arial", size=9, color="666666"); ws2.cell(row=row_sep+2, column=3).border = THIN_BORDER

# 成新折扣法
ws2.cell(row=row_sep+3, column=1, value="成新折扣法折旧").font = BOLD_BLACK; ws2.cell(row=row_sep+3, column=1).border = THIN_BORDER
ws2.cell(row=row_sep+3, column=2, value="=ROUND(B5*(1-B9),0)").font = BLACK_FONT
ws2.cell(row=row_sep+3, column=2).number_format = '#,##0'; ws2.cell(row=row_sep+3, column=2).border = THIN_BORDER
ws2.cell(row=row_sep+3, column=3, value="D = C × (1-q)。q=成新率").font = Font(name="Arial", size=9, color="666666"); ws2.cell(row=row_sep+3, column=3).border = THIN_BORDER

# 分解法
ws2.cell(row=row_sep+4, column=1, value="分解法 — 物质折旧").font = BOLD_BLACK; ws2.cell(row=row_sep+4, column=1).border = THIN_BORDER
ws2.cell(row=row_sep+4, column=2).font = BLUE_FONT; ws2.cell(row=row_sep+4, column=2).number_format = '#,##0'; ws2.cell(row=row_sep+4, column=2).border = THIN_BORDER
ws2.cell(row=row_sep+4, column=3, value="[原因: 结构老化/设备损坏/装修陈旧等]").font = Font(name="Arial", size=9, color="666666"); ws2.cell(row=row_sep+4, column=3).border = THIN_BORDER

ws2.cell(row=row_sep+5, column=1, value="分解法 — 功能折旧").font = BOLD_BLACK; ws2.cell(row=row_sep+5, column=1).border = THIN_BORDER
ws2.cell(row=row_sep+5, column=2).font = BLUE_FONT; ws2.cell(row=row_sep+5, column=2).number_format = '#,##0'; ws2.cell(row=row_sep+5, column=2).border = THIN_BORDER
ws2.cell(row=row_sep+5, column=3, value="[原因: 户型过时/层高低/无电梯等]").font = Font(name="Arial", size=9, color="666666"); ws2.cell(row=row_sep+5, column=3).border = THIN_BORDER

ws2.cell(row=row_sep+6, column=1, value="分解法 — 外部折旧").font = BOLD_BLACK; ws2.cell(row=row_sep+6, column=1).border = THIN_BORDER
ws2.cell(row=row_sep+6, column=2).font = BLUE_FONT; ws2.cell(row=row_sep+6, column=2).number_format = '#,##0'; ws2.cell(row=row_sep+6, column=2).border = THIN_BORDER
ws2.cell(row=row_sep+6, column=3, value="[原因: 污染/规划变更/周边衰退等]").font = Font(name="Arial", size=9, color="666666"); ws2.cell(row=row_sep+6, column=3).border = THIN_BORDER

ws2.cell(row=row_sep+6, column=2).comment = Comment("4.4.10-4.4.12条: 物质折旧(自然/使用/环境)、功能折旧(技术落后/设计缺陷)、外部折旧(区位/外部因素)", "GB/T 50291-2015")

# 折旧合计
row_dep_total = row_sep + 8
ws2.cell(row=row_dep_total, column=1, value="折旧合计 (元/m²)").font = Font(name="Arial", size=12, color="000000", bold=True)
ws2.cell(row=row_dep_total, column=1).border = THIN_BORDER

formula_dep = f'=ROUND(IF(B4="年龄-寿命法",B{row_sep+2},IF(B4="成新折扣法",B{row_sep+3},B{row_sep+4}+B{row_sep+5}+B{row_sep+6})),0)'
ws2.cell(row=row_dep_total, column=2, value=formula_dep).font = Font(name="Arial", size=12, color="000000", bold=True)
ws2.cell(row=row_dep_total, column=2).number_format = '#,##0'; ws2.cell(row=row_dep_total, column=2).fill = YELLOW_FILL
ws2.cell(row=row_dep_total, column=2).border = THIN_BORDER
ws2.cell(row=row_dep_total, column=3, value="IF 自动切换三种方法").font = Font(name="Arial", size=9, color="666666")

ws2.cell(row=row_dep_total, column=2).comment = Comment("按B4选择的折旧方法自动取对应结果。年龄-寿命法=直线法、成新折扣法=(1-q)×C、分解法=三个子项之和。", "公式")

# 数据验证
for row in [6, 7, 8]:
    dv_pos_int = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0")
    ws2.add_data_validation(dv_pos_int); dv_pos_int.add(ws2.cell(row=row, column=2))

dv_rate_q = DataValidation(type="decimal", operator="between", formula1="0", formula2="1")
ws2.add_data_validation(dv_rate_q); dv_rate_q.add(ws2['B9'])

# ═══════════════════════════════════════════════
# SHEET 3: 成本价值
# ═══════════════════════════════════════════════
ws3 = wb.create_sheet("成本价值")
ws3.column_dimensions['A'].width = 28
ws3.column_dimensions['B'].width = 22

# 标题
ws3.merge_cells('A1:B1')
c = ws3['A1']; c.value = "成本价值"; c.font = TITLE_FONT; c.fill = DARK_BLUE_FILL; c.alignment = CENTER
ws3.row_dimensions[1].height = 30

for i, h in enumerate(["项目", "金额"], 1):
    cell = ws3.cell(row=3, column=i, value=h)
    cell.font = Font(name="Arial", size=11, color="000000", bold=True)
    cell.fill = LIGHT_BLUE_FILL; cell.alignment = CENTER; cell.border = THIN_BORDER

s3_data = [
    (4, "重置成本合计 (元/m²)",  "=重置成本!C14",      "跨表引用"),
    (5, "建筑物折旧合计 (元/m²)", "=折旧测算!B19",     "跨表引用"),
    (6, "权益状况调整 (元/m²)",   "BLUE",              "如有租赁/抵押/查封等权益限制"),
    (7, "成本价值 (元/m²)",       "=ROUND(B4-B5+B6,0)",         "公式; 可加黄色底色"),
    (9, "建筑总面积 (m²)",        "BLUE",              "输入估价对象建筑面积"),
    (10, "成本总价值 (元)",       "=ROUND(B7*B9,0)",            "最终结果"),
]

for row, label, val, note in s3_data:
    ws3.cell(row=row, column=1, value=label).font = BOLD_BLACK; ws3.cell(row=row, column=1).border = THIN_BORDER
    c2 = ws3.cell(row=row, column=2)
    if val == "BLUE":
        c2.font = BLUE_FONT; c2.number_format = '#,##0'
    else:
        c2.font = BLACK_FONT; c2.value = val; c2.number_format = '#,##0'
    c2.border = THIN_BORDER; c2.alignment = CENTER

# 高亮关键结果
ws3['B7'].fill = YELLOW_FILL; ws3['B10'].fill = YELLOW_FILL

# 注释
ws3['B7'].comment = Comment("4.4.15条: 成本价值 = 重置成本 - 建筑物折旧 ± 权益状况调整。", "GB/T 50291-2015")

# Sheet 保护
for ws in [ws1, ws2, ws3]:
    ws.protection.sheet = True
    ws.protection.password = ""

out_path = os.path.join(OUTPUT_DIR, "成本法测算_模板.xlsx")
save_frozen(wb, out_path)
print(f"✅ 成本法模板已生成: {out_path}")
print(f"   Sheet 1: 重置成本 (7项支出 + 5个参数)")
print(f"   Sheet 2: 折旧测算 (3种方法 × IF自动切换)")
print(f"   Sheet 3: 成本价值 (跨表引用汇总)")
