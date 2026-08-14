#!/usr/bin/env python3
"""
房地产评估明细表 → 计算模板模糊化脚本

策略：
1. 保留所有公式、结构、表头、修正系数逻辑
2. 模糊化规则从 anonymize_rules.yaml 加载（与脚本同目录）
3. 输出到 outputs/ 目录

用法：
  python scripts/anonymize_template.py <源文件路径>

规则配置：
  编辑 scripts/anonymize_rules.yaml 修改替换规则，无需改代码。
"""

import re
import sys
import yaml
import openpyxl
from pathlib import Path

# 路径
SCRIPT_DIR = Path(__file__).parent
RULES_PATH = SCRIPT_DIR / "anonymize_rules.yaml"
DST = SCRIPT_DIR.parent / "outputs" / "房地产评估明细表-计算模板.xlsx"


def load_rules(rules_path: Path = RULES_PATH) -> dict:
    """从 YAML 加载模糊化规则。

    返回:
      {
        "text_replacements": dict[str, str],
        "regex_replacements": list[(compiled_pattern, replacement)],
        "numeric_replacements": dict[(sheet, coord), value],
        "verify_patterns": list[(pattern_str, description)],
      }
    """
    with open(rules_path, encoding="utf-8") as f:
        cfg = yaml.safe_load(f)

    # 文本替换：直接取 dict
    text_replacements = dict(cfg.get("text_replacements", {}))

    # 正则替换：编译 pattern
    regex_replacements = []
    for item in cfg.get("regex_replacements", []):
        pattern = re.compile(item["pattern"])
        regex_replacements.append((pattern, item["replacement"]))

    # 数值替换：把 "sheet!coord" → (sheet, coord)
    numeric_replacements = {}
    for key, value in cfg.get("numeric_replacements", {}).items():
        sheet, coord = key.rsplit("!", 1)
        numeric_replacements[(sheet, coord)] = value

    # 验证扫描
    verify_patterns = []
    for item in cfg.get("verify_patterns", []):
        verify_patterns.append((item["pattern"], item["description"]))

    return {
        "text_replacements": text_replacements,
        "regex_replacements": regex_replacements,
        "numeric_replacements": numeric_replacements,
        "verify_patterns": verify_patterns,
    }


def anonymize_value(value, rules):
    """对单元格值进行模糊化处理。"""
    if value is None:
        return None

    if isinstance(value, str):
        result = value
        # 跳过公式
        if result.startswith("="):
            return result
        # 精确替换
        for old, new in rules["text_replacements"].items():
            result = result.replace(old, new)
        # 正则替换
        for pattern, replacement in rules["regex_replacements"]:
            result = pattern.sub(replacement, result)
        return result if result.strip() else None

    return value


def process_workbook(src_path, rules):
    """处理整个工作簿。"""
    wb = openpyxl.load_workbook(src_path)
    numeric = rules["numeric_replacements"]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue

                coord = (sheet_name, cell.coordinate)

                # 1. 数值替换（精确单元格定位）
                if coord in numeric:
                    cell.value = numeric[coord]
                    continue

                # 2. 文本/正则模糊化
                cell.value = anonymize_value(cell.value, rules)

    # 保存
    DST.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(DST))
    wb.close()
    print(f"✅ 模糊化完成: {DST}")


def verify_anonymization(rules):
    """验证模糊化结果，扫描是否遗漏敏感信息。"""
    wb = openpyxl.load_workbook(str(DST))

    issues = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    for pattern, desc in rules["verify_patterns"]:
                        if pattern in cell.value:
                            issues.append(
                                f"  ⚠️ {sheet_name}!{cell.coordinate}: 发现{desc} '{pattern}'"
                            )

    wb.close()

    if issues:
        print("⚠️ 以下内容需人工确认：")
        for issue in issues:
            print(issue)
        return False
    else:
        print("✅ 验证通过：未发现明显敏感信息")
        return True


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python scripts/anonymize_template.py <源文件路径>")
        print(f"规则配置: {RULES_PATH}")
        sys.exit(1)

    rules = load_rules()
    print(f"加载规则: {RULES_PATH}")
    print(f"  文本替换: {len(rules['text_replacements'])} 条")
    print(f"  正则替换: {len(rules['regex_replacements'])} 条")
    print(f"  数值替换: {len(rules['numeric_replacements'])} 条")

    process_workbook(sys.argv[1], rules)
    print()
    verify_anonymization(rules)
