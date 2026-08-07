#!/usr/bin/env python3
"""生成比较法测算 Excel 模板 — GB/T 50291-2015 第4.2节"""
import os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

# ── 颜色定义 ──
BLUE_FONT = Font(name="Arial", size=11, color="0000FF")
BLACK_FONT = Font(name="Arial", size=11, color="000000")
BOLD_BLACK = Font(name="Arial", size=11, color="000000", bold=True)
HEADER_FONT = Font(name="Arial", size=12, color="FFFFFF", bold=True)
TITLE_FONT = Font(name="Arial", size=14, color="FFFFFF", bold=True)
LEGEND_FONT = Font(name="Arial", size=9, color="333333", italic=True)

DARK_BLUE_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "outputs", "templates")
os.makedirs(OUTPUT_DIR, exist_ok=True)

wb = Workbook()

# ═══════════════════════════════════════════════
# SHEET 1: 可比实例数据
# ═══════════════════════════════════════════════
ws1 = wb.active
ws1.title = "可比实例数据"

# 列宽
ws1.column_dimensions['A'].width = 22
for col in 'BCDEF':
    ws1.column_dimensions[col].width = 18

# 行1: 标题
ws1.merge_cells('A1:F1')
c = ws1['A1']
c.value = "比较法测算表"
c.font = TITLE_FONT
c.fill = DARK_BLUE_FILL
c.alignment = CENTER
ws1.row_dimensions[1].height = 30

# 行2: 项目信息
ws1.merge_cells('A2:B2')
ws1['A2'] = "估价对象: [请输入地址]"
ws1['A2'].font = BLUE_FONT; ws1['A2'].alignment = LEFT_WRAP
ws1.merge_cells('C2:D2')
ws1['C2'] = "价值时点: [请输入日期]"
ws1['C2'].font = BLUE_FONT; ws1['C2'].alignment = LEFT_WRAP
ws1.merge_cells('E2:F2')
ws1['E2'] = "单位: 元/m²"
ws1['E2'].font = BLACK_FONT; ws1['E2'].alignment = LEFT_WRAP

# 行3: 颜色图例
ws1.merge_cells('A3:F3')
ws1['A3'] = "图例: 蓝色字=可修改输入  |  黑色字=公式自动计算  |  黄色底=关键参数"
ws1['A3'].font = LEGEND_FONT; ws1['A3'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# 行5: 表头
headers = ["项目", "实例A", "实例B", "实例C", "实例D", "数据来源"]
for i, h in enumerate(headers, 1):
    cell = ws1.cell(row=5, column=i, value=h)
    cell.font = BOLD_BLACK; cell.fill = LIGHT_BLUE_FILL
    cell.alignment = CENTER; cell.border = THIN_BORDER
ws1.row_dimensions[5].height = 24

# 行6-14: 数据行
row_data = [
    (6,  "成交价格 (元/m²)",      "BLUE",  "如: 25000", "可比实例实际成交单价，需≥3个"),
    (7,  "标准化调整系数",         "BLUE",  "1.00",      "财产范围/付款方式/融资条件/税费负担/计价单位统一"),
    (8,  "交易情况修正系数",       "BLUE",  "1.00",      "正常交易=1.00; 异常交易据实调整; 范围0.80~1.20 (≤±20%)"),
    (9,  "市场状况调整系数",       "BLUE",  "1.00",      "成交日至价值时点价格指数比值; 距价值时点≤2年"),
    (10, "区位状况调整系数",       "BLUE",  "1.00",      "位置/交通/配套/环境/楼层/朝向; 范围0.80~1.20"),
    (11, "实物状况调整系数",       "BLUE",  "1.00",      "面积/结构/装修/新旧/布局; 范围0.80~1.20"),
    (12, "权益状况调整系数",       "BLUE",  "1.00",      "规划/年限/共有/租赁/查封; 范围0.80~1.20"),
]

for row, label, style, placeholder, comment_text in row_data:
    ws1.cell(row=row, column=1, value=label).font = BOLD_BLACK
    for col_idx in range(2, 6):
        cell = ws1.cell(row=row, column=col_idx)
        if style == "BLUE":
            cell.font = BLUE_FONT
            cell.value = placeholder if col_idx == 2 else None
        cell.border = THIN_BORDER
        cell.alignment = CENTER
    # 数据来源列
    src_cell = ws1.cell(row=row, column=6, value=comment_text)
    src_cell.font = Font(name="Arial", size=9, color="666666", italic=True)
    src_cell.alignment = LEFT_WRAP

# 行13: 比较价格 (公式)
ws1.cell(row=13, column=1, value="比较价格 (元/m²)").font = BOLD_BLACK
for col_idx in range(2, 6):
    col_letter = get_column_letter(col_idx)
    formula = f"=ROUND({col_letter}6*{col_letter}7*{col_letter}8*{col_letter}9*{col_letter}10*{col_letter}11*{col_letter}12,0)"
    cell = ws1.cell(row=13, column=col_idx, value=formula)
    cell.font = BLACK_FONT; cell.number_format = '#,##0'
    cell.border = THIN_BORDER; cell.alignment = CENTER
ws1.cell(row=13, column=6, value="公式: 成交价×五项修正系数连乘").font = Font(name="Arial", size=9, color="666666", italic=True)
ws1.row_dimensions[13].height = 22

# 行14: 综合修正幅度 (公式)
ws1.cell(row=14, column=1, value="综合修正幅度").font = BOLD_BLACK
for col_idx in range(2, 6):
    col_letter = get_column_letter(col_idx)
    formula = f"=ABS({col_letter}13/{col_letter}6-1)"
    cell = ws1.cell(row=14, column=col_idx, value=formula)
    cell.font = BLACK_FONT; cell.number_format = '0.0%'
    cell.border = THIN_BORDER; cell.alignment = CENTER
ws1.cell(row=14, column=6, value="红线≤30%").font = Font(name="Arial", size=9, color="FF0000", italic=True)
ws1.row_dimensions[14].height = 22

# ── 数据验证 (Sheet 1) ──
# 成交价格 > 0
dv_price = DataValidation(type="decimal", operator="greaterThan", formula1="0")
dv_price.error = "成交价格必须大于0"
dv_price.errorTitle = "输入错误"
ws1.add_data_validation(dv_price)
for col_idx in range(2, 6):
    dv_price.add(ws1.cell(row=6, column=col_idx))

# 修正系数 0.80~1.20
dv_factor = DataValidation(type="decimal", operator="between", formula1="0.80", formula2="1.20")
dv_factor.error = "修正系数应在0.80~1.20之间 (±20%)"
dv_factor.errorTitle = "红线约束"
ws1.add_data_validation(dv_factor)
for row in [7, 8, 9, 10, 11, 12]:
    for col_idx in range(2, 6):
        dv_factor.add(ws1.cell(row=row, column=col_idx))

# ── 单元格注释 (规范条文) ──
comment_map_s1 = {
    (6, 1): "4.2.3条: 可比实例≥3个。成交价格为正常成交价格或可修正为正常成交价格。",
    (7, 1): "4.2.10条: 建立比较基础，统一财产范围/付款方式/融资条件/税费负担/计价单位。",
    (8, 1): "4.2.11条: 交易情况修正。正常交易=100，非正常交易据实调整，单项≤±20%。",
    (9, 1): "4.2.12条: 市场状况调整。同质房地产价格指数、变动率调整，距价值时点≤2年。",
    (10, 1): "4.2.13条: 区位状况调整。位置/交通/配套/环境/楼层/朝向。",
    (11, 1): "4.2.13条: 实物状况调整。面积/结构/装修/新旧程度/布局。",
    (12, 1): "4.2.13条: 权益状况调整。规划/年限/共有/租赁/查封/地役权。",
    (13, 1): "公式: 比较价格 = 成交价格 × 各项修正系数连乘。",
    (14, 1): "4.2.15条第1款: 综合修正幅度≤30%，超标需说明。",
}
for pos, text in comment_map_s1.items():
    ws1.cell(row=pos[0], column=pos[1]).comment = Comment(text, "GB/T 50291-2015")

# ── 条件格式: 综合修正幅度 > 30% 变红 ──
for col_idx in range(2, 6):
    col_letter = get_column_letter(col_idx)
    ws1.conditional_formatting.add(
        f"{col_letter}14",
        CellIsRule(operator="greaterThan", formula=["0.30"],
                   font=Font(color="FF0000", bold=True), fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )

# ── Sheet 保护 ──
ws1.protection.sheet = True
ws1.protection.password = ""
# 解锁蓝色输入单元格 (行6-12, 列B-E)
from openpyxl.styles import Protection
for r in range(6, 13):
    for c in range(2, 6):
        cell = ws1.cell(row=r, column=c)
        cell.protection = Protection(locked=False)
# 解锁项目信息行
ws1['A2'].protection = Protection(locked=False)
ws1['C2'].protection = Protection(locked=False)

# ═══════════════════════════════════════════════
# SHEET 2: 比较价值汇总
# ═══════════════════════════════════════════════
ws2 = wb.create_sheet("比较价值汇总")
ws2.column_dimensions['A'].width = 28
ws2.column_dimensions['B'].width = 22
ws2.column_dimensions['C'].width = 18

# 行1: 标题
ws2.merge_cells('A1:C1')
c = ws2['A1']
c.value = "比较价值汇总"
c.font = TITLE_FONT; c.fill = DARK_BLUE_FILL; c.alignment = CENTER
ws2.row_dimensions[1].height = 30

# 行3: 列标题
for i, h in enumerate(["指标", "数值", "红线"], 1):
    cell = ws2.cell(row=3, column=i, value=h)
    cell.font = BOLD_BLACK; cell.fill = LIGHT_BLUE_FILL
    cell.alignment = CENTER; cell.border = THIN_BORDER

# 数据行
s2_data = [
    (4, "比较价值-算术平均 (元/m²)", "=ROUND(AVERAGE(可比实例数据!B13:E13),0)", "—"),
    (5, "比较价值-加权平均 (元/m²)", "=ROUND(SUMPRODUCT(可比实例数据!B13:E13*{1,1,1,1})/SUMPRODUCT({1,1,1,1}),0)", "—"),
    (6, "比较价值-中位数 (元/m²)", "=ROUND(MEDIAN(可比实例数据!B13:E13),0)", "—"),
    (8, "权重 A/B/C/D", "YELLOW_INPUT", "合计100%"),
    (10, "最高价/最低价 极差比", "=IF(COUNT(可比实例数据!B13:E13)>=2,MAX(可比实例数据!B13:E13)/MIN(可比实例数据!B13:E13),\"需≥2个实例\")", "≤1.2"),
    (11, "综合修正幅度-最大", "=MAX(可比实例数据!B14:E14)", "≤30%"),
    (12, "单项修正幅度-最大", "=MAX(ABS(可比实例数据!B8-1),ABS(可比实例数据!B9-1),ABS(可比实例数据!B10-1),ABS(可比实例数据!B11-1),ABS(可比实例数据!B12-1))", "≤20%"),
]
for row, label, val, constraint in s2_data:
    ws2.cell(row=row, column=1, value=label).font = BOLD_BLACK
    c1 = ws2.cell(row=row, column=1); c1.border = THIN_BORDER

    if val == "YELLOW_INPUT":
        cell = ws2.cell(row=row, column=2)
        cell.font = BLUE_FONT; cell.value = None
        cell.fill = YELLOW_FILL
        cell.alignment = CENTER
        cell.comment = Comment("输入各可比实例权重(百分比)，合计应为100%", "GB/T 50291-2015")
    else:
        cell = ws2.cell(row=row, column=2, value=val)
        cell.font = BLACK_FONT; cell.number_format = '#,##0'
        cell.alignment = CENTER
    cell.border = THIN_BORDER

    c3 = ws2.cell(row=row, column=3, value=constraint)
    c3.font = Font(name="Arial", size=9, color="FF0000" if constraint.startswith("≤") else "666666")
    c3.alignment = CENTER; c3.border = THIN_BORDER

# 行8权重行4列输入
ws2.cell(row=8, column=2, value=None).value = None
ws2.cell(row=8, column=2).font = BLUE_FONT

# 条件格式: 极差比 > 1.2 红色
ws2.conditional_formatting.add("B10", CellIsRule(operator="greaterThan", formula=["1.2"],
    font=Font(color="FF0000", bold=True), fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")))
ws2.conditional_formatting.add("B11", CellIsRule(operator="greaterThan", formula=["0.30"],
    font=Font(color="FF0000", bold=True), fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")))
ws2.conditional_formatting.add("B12", CellIsRule(operator="greaterThan", formula=["0.20"],
    font=Font(color="FF0000", bold=True), fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")))

# 注释
ws2.cell(row=4, column=1).comment = Comment("简单算术平均，适用于各实例可比性相近。", "valuation")
ws2.cell(row=6, column=1).comment = Comment("中位数不受极端值影响，推荐用作主要参考。", "valuation")
ws2.cell(row=10, column=1).comment = Comment("4.2.15条第3款: 各可比实例经修正调整后的价格比值≤1.2。", "GB/T 50291-2015")

# Sheet 保护
ws2.protection.sheet = True
ws2.protection.password = ""

# ── 保存 ──
out_path = os.path.join(OUTPUT_DIR, "比较法测算_模板.xlsx")
wb.save(out_path)
print(f"✅ 比较法模板已生成: {out_path}")
print(f"   Sheet 1: 可比实例数据 (19行)")
print(f"   Sheet 2: 比较价值汇总 (13行)")
print(f"   数据验证: 成交价格>0, 修正系数0.80~1.20")
print(f"   条件格式: 综合修正>30%红色, 极差>1.2红色")
