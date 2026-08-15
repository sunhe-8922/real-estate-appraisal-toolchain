#!/usr/bin/env python3
"""
populate_excel_from_schema.py — Schema → Excel 数据回填与验证

工作流：
  1. 从 JSON schema 读取输入字段（实例单价、修正系数、面积等）
  2. 写入 Excel 模板对应单元格（REVERSE_LOOKUP 反向映射）
  3. 读取 Excel 公式计算结果
  4. 与 Schema 已有输出字段交叉验证
  5. 可选：把验证通过的 Excel 结果写回 Schema（回填模式）

用法：
  # 填入并输出填充后的 Excel
  python scripts/populate_excel_from_schema.py \\
      --schema schema/example-武汉洪山住宅.json \\
      --template outputs/房地产评估明细表-计算模板.xlsx \\
      --output outputs/result-populated.xlsx

  # 仅验证（不修改文件）
  python scripts/populate_excel_from_schema.py \\
      --schema schema/example-武汉洪山住宅.json \\
      --template outputs/房地产评估明细表-计算模板.xlsx \\
      --verify-only

  # 回填模式：Excel 结果 → Schema JSON
  python scripts/populate_excel_from_schema.py \\
      --schema schema/example-武汉洪山住宅.json \\
      --template outputs/房地产评估明细表-计算模板.xlsx \\
      --output outputs/result-populated.json \\
      --backfill
"""

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any, Optional

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# ── REVERSE_LOOKUP 镜像（与 rebuild_excel_formula.py 保持一致）──────────
INSTANCE_COLS = {0: ("T", "V"), 1: ("W", "Y"), 2: ("Z", "AB")}

REVERSE_MAP = [
    # 比较法实例
    ("methods.comps.comparableInstances[{i}].unitPrice", "市场价比较法", "{eval_col}4"),
    ("methods.comps.comparableInstances[{i}].adjustments.transactionSituation", "市场价比较法", "{eval_col}5"),
    ("methods.comps.comparableInstances[{i}].adjustments.marketCondition", "市场价比较法", "{eval_col}6"),
    ("methods.comps.comparableInstances[{i}].adjustments.locationDetails", "市场价比较法", "{inst_col}7:{inst_col}18"),
    ("methods.comps.comparableInstances[{i}].adjustments.interestDetails", "市场价比较法", "{inst_col}19:{inst_col}23"),
    ("methods.comps.comparableInstances[{i}].adjustments.physicalDetails", "市场价比较法", "{inst_col}24:{inst_col}31"),
    ("methods.comps.comparableInstances[{i}].adjustedUnitPrice", "市场价比较法", "{eval_col}32"),
    # 收益法
    ("methods.income.netOperatingIncome.annualAmount", "住宅-收益法测算", "G4"),
    ("methods.income.netOperatingIncome.effectiveGrossIncome", "住宅-收益法测算", "G5"),
    ("methods.income.netOperatingIncome.operatingExpenses", "住宅-收益法测算", "G11"),
    ("methods.income.netOperatingIncome.growthRate", "住宅-收益法测算", "G25"),
    ("methods.income.rate.value", "住宅-收益法测算", "G24"),
    ("methods.income.holdingPeriod", "住宅-收益法测算", "G26"),
    ("methods.income.finalValue.total", "住宅-收益法测算", "G27"),
    ("result.weightAllocation.income", "住宅-收益法测算", "I27"),
    ("result.weightAllocation.comps", "住宅-收益法测算", "I28"),
    ("methods.comps.finalValue.total", "住宅-收益法测算", "J28"),
    ("result.finalTotalValue", "住宅-收益法测算", "J29"),
    # 评估明细表
    ("property.area", "评估明细表", "M6"),
    ("result.finalUnitValue", "评估明细表", "N6"),
    ("result.finalTotalValue", "评估明细表", "O6"),
]


def _build_reverse_lookup() -> dict[str, str]:
    lookup = {}
    for path, sheet, pattern in REVERSE_MAP:
        if "{i}" in path:
            for i, (eval_col, inst_col) in INSTANCE_COLS.items():
                p = path.replace("{i}", str(i))
                cell = f"{sheet}!{pattern.format(eval_col=eval_col, inst_col=inst_col)}"
                lookup[p] = cell
        else:
            lookup[path] = f"{sheet}!{pattern}"
    return lookup


REVERSE_LOOKUP = _build_reverse_lookup()


def get_nested(data: dict, path: str) -> Any:
    """根据点分路径获取嵌套字典值。支持 arrays[index] 语法。"""
    current = data
    for key in re.findall(r'[\w]+|\[\d+\]', path):
        key = key.replace('[', '').replace(']', '')
        if isinstance(current, dict):
            if key in current:
                current = current[key]
            else:
                return None
        elif isinstance(current, list):
            try:
                idx = int(key)
                current = current[idx]
            except (ValueError, IndexError):
                return None
        else:
            return None
    return current


def set_nested(data: dict, path: str, value: Any) -> bool:
    """根据点分路径设置嵌套字典值。"""
    keys = re.findall(r'[\w]+|\[\d+\]', path)
    current = data
    for key in keys[:-1]:
        key = key.replace('[', '').replace(']', '')
        if isinstance(current, dict):
            if key in current:
                current = current[key]
            else:
                return False
        elif isinstance(current, list):
            try:
                idx = int(key)
                current = current[idx]
            except (ValueError, IndexError):
                return False
        else:
            return False
    last_key = keys[-1].replace('[', '').replace(']', '')
    if isinstance(current, dict):
        current[last_key] = value
        return True
    elif isinstance(current, list):
        try:
            idx = int(last_key)
            current[idx] = value
            return True
        except (ValueError, IndexError):
            return False
    return False


def read_cell(wb, cell_ref: str) -> Any:
    """读取 Excel 单元格值（data_only=True）。"""
    m = re.match(r'^([^\[]+)(?:\[([^\]]+)\])?!(.+)$', cell_ref)
    if not m:
        return None
    sheet_name = m.group(1)
    cell_spec = m.group(3)  # e.g. "T32"
    m2 = re.match(r'^([A-Z]+)(\d+)$', cell_spec)
    if not m2:
        return None
    col, row = m2.group(1), int(m2.group(2))
    ws = wb[sheet_name]
    if ws is None:
        return None
    cell = ws.cell(row=row, column=openpyxl.utils.column_index_from_string(col))
    return cell.value


def write_cell(ws, col: str, row: int, value: Any) -> None:
    """写入 Excel 单元格。"""
    cell = ws.cell(row=row, column=openpyxl.utils.column_index_from_string(col))
    cell.value = value


def write_array_range(ws, col: str, start_row: int, values: list) -> None:
    """写入数组范围（locationDetails 等子项列表）。"""
    for i, v in enumerate(values):
        cell = ws.cell(row=start_row + i, column=openpyxl.utils.column_index_from_string(col))
        cell.value = v


def extract_cell_ref(cell_ref: str) -> tuple[str, str, int]:
    """解析 'Sheet!ColRow' → (sheet_name, col, row)。"""
    m = re.match(r'^([^\[]+)(?:\[([^\]]+)\])?!(.+)$', cell_ref)
    if not m:
        return None, None, None
    sheet_name = m.group(1)
    cell_spec = m.group(3)
    m2 = re.match(r'^([A-Z]+)(\d+)$', cell_spec)
    if not m2:
        return None, None, None
    return sheet_name, m2.group(1), int(m2.group(2))


def populate_excel(schema_path: str, template_path: str, output_path: Optional[str] = None,
                   verify_only: bool = False, backfill: bool = False) -> dict:
    """主回填流程。"""
    with open(schema_path, encoding='utf-8') as f:
        schema_data = json.load(f)

    wb = openpyxl.load_workbook(template_path, data_only=False)
    results = {"populate": [], "verify": [], "backfill": []}

    # ── 阶段1：Schema → Excel（填入输入数据）────────────────────────────
    for json_path, cell_ref in REVERSE_LOOKUP.items():
        if not cell_ref or ':' in cell_ref:
            continue  # 跳过数组范围（单独处理）

        val = get_nested(schema_data, json_path)
        if val is None:
            continue

        sheet_name, col, row = extract_cell_ref(cell_ref)
        if not sheet_name or not col:
            continue

        ws = wb[sheet_name]
        if ws:
            write_cell(ws, col, row, val)
            results["populate"].append({
                "json_path": json_path,
                "excel": cell_ref,
                "value": val
            })

    # 处理数组范围（locationDetails 等）
    array_paths = [p for p in REVERSE_LOOKUP if ':' in REVERSE_LOOKUP[p]]
    for json_path in array_paths:
        cell_ref = REVERSE_LOOKUP[json_path]
        sheet_name, col_start, start_row = extract_cell_ref(cell_ref.split(':')[0])
        _, col_end, end_row = extract_cell_ref(cell_ref.split(':')[1])
        if not sheet_name or not col_start or not col_end:
            continue

        val_list = get_nested(schema_data, json_path)
        if not isinstance(val_list, list):
            continue

        ws = wb[sheet_name]
        if ws:
            # 写入每个子项的 factor
            for i, item in enumerate(val_list):
                factor = item.get("factor") if isinstance(item, dict) else item
                if factor is not None and (start_row + i) <= end_row:
                    write_cell(ws, col_start, start_row + i, factor)
                    results["populate"].append({
                        "json_path": f"{json_path}[{i}].factor",
                        "excel": f"{sheet_name}!{col_start}{start_row + i}",
                        "value": factor
                    })

    # ── 阶段2：保存 + 验证 ───────────────────────────────────────────────
    excel_output = None
    if not verify_only:
        # 确定 Excel 输出路径（即使 --backfill 也要先写 Excel）
        if output_path and output_path.endswith('.json'):
            # --backfill 模式：输出 JSON，但中间 Excel 存临时路径
            excel_output = (Path(output_path).parent / (Path(output_path).stem + "-filled.xlsx")).__str__()
        else:
            excel_output = output_path or template_path.replace('.xlsx', '-filled.xlsx')
        wb.save(excel_output)
        wb.close()
        # 用 data_only=True 重新打开读取计算值
        wb = openpyxl.load_workbook(excel_output, data_only=True)

    # 读取计算结果并验证
    for json_path, cell_ref in REVERSE_LOOKUP.items():
        if not cell_ref or ':' in cell_ref:
            continue
        excel_val = read_cell(wb, cell_ref)
        schema_val = get_nested(schema_data, json_path)

        if excel_val is not None and schema_val is not None:
            try:
                diff = abs(float(excel_val) - float(schema_val))
            except (ValueError, TypeError):
                diff = None
            status = "✅" if (diff is None or diff < 1) else "🚨"
            results["verify"].append({
                "json_path": json_path,
                "excel": cell_ref,
                "excel_value": excel_val,
                "schema_value": schema_val,
                "diff": diff,
                "status": status
            })

    # ── 阶段3：Excel → Schema（回填）────────────────────────────────────
    if backfill:
        for item in results["verify"]:
            if item["status"] == "✅":
                new_val = item["excel_value"]
                if set_nested(schema_data, item["json_path"], new_val):
                    results["backfill"].append({
                        "json_path": item["json_path"],
                        "new_value": new_val
                    })
        # 保存回填后的 Schema
        if output_path and output_path.endswith('.json'):
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(schema_data, f, ensure_ascii=False, indent=2)
            print(f"\n✓ 回填结果已保存到: {output_path}")
        elif not verify_only and excel_output:
            # 同时保存回填后的 Schema 到同目录
            schema_out = (Path(excel_output).parent / (Path(excel_output).stem + ".json")).__str__()
            with open(schema_out, 'w', encoding='utf-8') as f:
                json.dump(schema_data, f, ensure_ascii=False, indent=2)
            print(f"\n✓ 回填结果已保存到: {schema_out}")

    wb.close()
    return results


def main():
    parser = argparse.ArgumentParser(description="Schema ↔ Excel 双向数据同步")
    parser.add_argument("--schema", required=True, help="Schema JSON 文件路径")
    parser.add_argument("--template", required=True, help="Excel 模板路径")
    parser.add_argument("--output", help="输出文件路径（Excel 或 JSON）")
    parser.add_argument("--verify-only", action="store_true", help="仅验证，不写入文件")
    parser.add_argument("--backfill", action="store_true", help="将 Excel 计算结果回填到 Schema")
    args = parser.parse_args()

    results = populate_excel(
        schema_path=args.schema,
        template_path=args.template,
        output_path=args.output,
        verify_only=args.verify_only,
        backfill=args.backfill
    )

    print("=" * 70)
    print("阶段1: Schema → Excel 填入数据")
    for item in results["populate"]:
        print(f"  {item['json_path'][:55]:55s} → {item['excel']:20s} = {item['value']}")

    print("\n阶段2: Excel 计算结果验证")
    for item in results["verify"]:
        print(f"  {item['status']} {item['json_path'][:40]:40s} | Excel={item['excel_value']} | Schema={item['schema_value']} | diff={item['diff']}")

    if results["backfill"]:
        print("\n阶段3: Excel → Schema 回填")
        for item in results["backfill"]:
            print(f"  {item['json_path'][:55]} = {item['new_value']}")


if __name__ == "__main__":
    main()
