#!/usr/bin/env python3
"""生成收益法测算 Excel 模板 — GB/T 50291-2015 第4.3节"""
import os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

# ── 颜色定义 ──
BLUE_FONT = Font(name="Arial", size=11, color="0000FF")
BLACK_FONT = Font(name="Arial", size=11, color="000000")
BOLD_BLACK = Font(name="Arial", size=11, color="000000", bold=True)
HEADER_FONT = Font(name="Arial", size=12, color="FFFFFF", bold=True)
TITLE_FONT = Font(name="Arial", size=14, color="FFFFFF", bold=True)
LEGEND_FONT = Font(name="Arial", size=9, color="333333", italic=True)

DARK_BLUE_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "outputs", "templates")
os.makedirs(OUTPUT_DIR, exist_ok=True)

wb = Workbook()

# ═══════════════════════════════════════════════
# SHEET 1: 净收益测算
# ═══════════════════════════════════════════════
ws1 = wb.active
ws1.title = "净收益测算"
ws1.column_dimensions['A'].width = 30
ws1.column_dimensions['B'].width = 20
ws1.column_dimensions['C'].width = 35

# 标题行
ws1.merge_cells('A1:C1')
c = ws1['A1']; c.value = "净收益测算"; c.font = TITLE_FONT; c.fill = DARK_BLUE_FILL; c.alignment = CENTER
ws1.row_dimensions[1].height = 30

# 项目信息
ws1.merge_cells('A2:B2')
ws1['A2'] = "估价对象: [请输入地址]"; ws1['A2'].font = BLUE_FONT
ws1['C2'] = "价值时点: [请输入日期]"; ws1['C2'].font = BLUE_FONT

# 颜色图例
ws1.merge_cells('A3:C3')
ws1['A3'] = "图例: 蓝色字=可修改输入  |  黑色字=公式自动计算  |  黄色底=关键参数"
ws1['A3'].font = LEGEND_FONT; ws1['A3'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# 表头
for i, h in enumerate(["项目", "数值", "数据来源"], 1):
    cell = ws1.cell(row=5, column=i, value=h)
    cell.font = BOLD_BLACK; cell.fill = LIGHT_BLUE_FILL; cell.alignment = CENTER; cell.border = THIN_BORDER

# 数据行
s1_data = [
    (6,  "潜在毛租金收入 (元/m²·年)",   "BLUE", 600,    "4.3.7条: 租赁收入法; 来源: 市场调查/合同"),
    (7,  "空置和收租损失率",           "BLUE_PCT", 0.05, "4.3.8条; 区域空置率调查"),
    (8,  "有效毛收入 (元/m²·年)",     "FORMULA", '=ROUND(B6*(1-B7),0)', "公式"),
    (9,  "其他收入 (元/m²·年)",       "BLUE", 50,     "押金利息/停车费/广告位等"),
    (10, "房地产税",                   "BLUE", None,   "城市房地产税/城镇土地使用税"),
    (11, "房屋保险费",                 "BLUE", None,   "建筑物及设备保险费"),
    (12, "物业服务费",                 "BLUE", None,   "物业管理公司收费"),
    (13, "管理费用",                   "BLUE", None,   "管理人员工资/办公费等"),
    (14, "维修费",                     "BLUE", None,   "日常维修和大修基金"),
    (15, "水电费",                     "BLUE", None,   "公共区域水电"),
    (16, "其他运营费用",               "BLUE", None,   "其他正常运营必要支出"),
    (17, "运营费用合计",               "FORMULA", '=ROUND(SUM(B10:B16),0)', "公式"),
    (18, "净收益 NOI (元/m²·年)",     "FORMULA", '=ROUND(B8+B9-B17,0)', "4.3.9条: 净收益 = 有效毛收入 - 运营费用"),
    (19, "最近三年实际净收益: Y-2",   "BLUE", None,   "4.3.11条: 必须调查近三年实际数据"),
    (20, "最近三年实际净收益: Y-1",   "BLUE", None,   "与预测值差异需分析修正"),
    (21, "最近三年实际净收益: Y-0",   "BLUE", None,   "以正常客观值为准"),
]

for row, label, style, default, comment_text in s1_data:
    c1 = ws1.cell(row=row, column=1, value=label)
    c1.font = BOLD_BLACK; c1.border = THIN_BORDER

    c2 = ws1.cell(row=row, column=2)
    if style == "BLUE":
        c2.font = BLUE_FONT; c2.value = default if default != "0" else None
        c2.number_format = '#,##0'
    elif style == "BLUE_PCT":
        c2.font = BLUE_FONT; c2.value = default
        c2.number_format = '0.0%'
    elif style == "FORMULA":
        c2.font = BLACK_FONT; c2.value = default
        c2.number_format = '#,##0'
    c2.border = THIN_BORDER; c2.alignment = CENTER

    c3 = ws1.cell(row=row, column=3, value=comment_text)
    c3.font = Font(name="Arial", size=9, color="666666", italic=True)
    c3.alignment = LEFT_WRAP; c3.border = THIN_BORDER

# 黄底标记关键行
ws1['B7'].fill = YELLOW_FILL  # 空置率
ws1['B18'].fill = YELLOW_FILL  # 净收益

# 数据验证
dv_pct = DataValidation(type="decimal", operator="between", formula1="0", formula2="1")
dv_pct.error = "空置率应在0~100%之间"; dv_pct.errorTitle = "输入错误"
ws1.add_data_validation(dv_pct); dv_pct.add(ws1['B7'])

dv_positive = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0")
dv_positive.error = "数值不能为负"; ws1.add_data_validation(dv_positive)
for row in [6, 9, 10, 11, 12, 13, 14, 15, 16, 19, 20, 21]:
    dv_positive.add(ws1.cell(row=row, column=2))

# 单元格注释
ws1['B6'].comment = Comment("4.3.7条: 潜在毛租金收入 = 可出租面积 × 市场租金水平。应调查周边同类物业实际租金。", "GB/T 50291-2015")
ws1['B7'].comment = Comment("4.3.8条: 空置和收租损失按同类物业正常客观水平确定。", "GB/T 50291-2015")
ws1['B18'].comment = Comment("4.3.9条: 净收益 = 有效毛收入 - 运营费用。各项收入/费用均取正常客观值(4.3.10条)。", "GB/T 50291-2015")
ws1['B19'].comment = Comment("4.3.11条: 必须调查近三年实际收入/费用数据，作为预测验证依据。", "GB/T 50291-2015")

# ═══════════════════════════════════════════════
# SHEET 2: 收益价值
# ═══════════════════════════════════════════════
ws2 = wb.create_sheet("收益价值")

# 列宽
for col_idx, w in enumerate([8, 16, 14, 16, 22], 1):
    ws2.column_dimensions[get_column_letter(col_idx)].width = w

# 标题
ws2.merge_cells('A1:E1')
c = ws2['A1']; c.value = "收益价值计算"; c.font = TITLE_FONT; c.fill = DARK_BLUE_FILL; c.alignment = CENTER
ws2.row_dimensions[1].height = 30

# 参数区
ws2.merge_cells('A2:B2')
ws2['A2'] = "报酬率 Y"; ws2['A2'].font = BOLD_BLACK; ws2['A2'].border = THIN_BORDER
ws2['C2'].font = BLUE_FONT; ws2['C2'].fill = YELLOW_FILL; ws2['C2'].number_format = '0.00%'
ws2['C2'].comment = Comment("报酬率确定方法: 市场提取法(≥3实例)/累加法/排序插入法。4.3.12-4.3.14条", "GB/T 50291-2015")

ws2.merge_cells('D2:E2')
ws2['D2'] = "方法: [市场提取/累加/排序]"; ws2['D2'].font = Font(name="Arial", size=9, color="666666")

ws2.merge_cells('A3:B3')
ws2['A3'] = "收益期 (年)"; ws2['A3'].font = BOLD_BLACK; ws2['A3'].border = THIN_BORDER
ws2['C3'].font = BLUE_FONT; ws2['C3'].fill = YELLOW_FILL; ws2['C3'].number_format = '0'
ws2['C3'].comment = Comment("全剩余寿命模式=收益期; 持有加转售模式=持有期。土地使用权届满则取两者较短者。", "GB/T 50291-2015")
ws2.merge_cells('D3:E3')
ws2['D3'] = "模式: [全剩余寿命/持有加转售]"; ws2['D3'].font = Font(name="Arial", size=9, color="666666")

# 表头
for i, h in enumerate(["期数", "净收益 (元/m²)", "折现系数", "现值 (元/m²)", "备注"], 1):
    cell = ws2.cell(row=5, column=i, value=h)
    cell.font = BOLD_BLACK; cell.fill = LIGHT_BLUE_FILL; cell.alignment = CENTER; cell.border = THIN_BORDER

# 生成15行折现行 (足够大多数场景)
MAX_YEARS = 60
for yr in range(1, MAX_YEARS + 1):
    row = 5 + yr
    # 期数
    ws2.cell(row=row, column=1, value=yr).font = BLACK_FONT
    ws2.cell(row=row, column=1).alignment = CENTER; ws2.cell(row=row, column=1).border = THIN_BORDER
    # 净收益 (跨表引用)
    formula_income = f"=ROUND(IF(A{row}<=$C$3, 净收益测算!B18, 0),0)"
    ws2.cell(row=row, column=2, value=formula_income).font = BLACK_FONT
    ws2.cell(row=row, column=2).number_format = '#,##0'; ws2.cell(row=row, column=2).border = THIN_BORDER
    # 折现系数
    formula_dcf = f"=IF(A{row}<=$C$3,1/(1+$C$2)^A{row},0)"
    ws2.cell(row=row, column=3, value=formula_dcf).font = BLACK_FONT
    ws2.cell(row=row, column=3).number_format = '0.000000'; ws2.cell(row=row, column=3).border = THIN_BORDER
    # 现值
    formula_pv = f"=ROUND(B{row}*C{row},0)"
    ws2.cell(row=row, column=4, value=formula_pv).font = BLACK_FONT
    ws2.cell(row=row, column=4).number_format = '#,##0'; ws2.cell(row=row, column=4).border = THIN_BORDER

# 收益价值汇总
row_val = 5 + MAX_YEARS + 1
ws2.merge_cells(f'A{row_val}:C{row_val}')
ws2.cell(row=row_val, column=1, value="收益价值 (全剩余寿命模式, 元/m²)").font = BOLD_BLACK
ws2.cell(row=row_val, column=1).alignment = LEFT_WRAP
formula_sum = f"=ROUND(SUM(D6:D{5+MAX_YEARS}),0)"
ws2.cell(row=row_val, column=4, value=formula_sum).font = Font(name="Arial", size=12, color="000000", bold=True)
ws2.cell(row=row_val, column=4).number_format = '#,##0'; ws2.cell(row=row_val, column=4).fill = YELLOW_FILL

# 期末转售收益区
row_tr = row_val + 2
ws2.merge_cells(f'A{row_tr}:B{row_tr}')
ws2.cell(row=row_tr, column=1, value="期末转售收益 (元/m²)").font = BOLD_BLACK
ws2.cell(row=row_tr, column=1).comment = Comment("仅持有加转售模式需填。期末转售收益不得用成本法测算(4.5.7条)", "GB/T 50291-2015")
ws2.cell(row=row_tr, column=3).font = BLUE_FONT; ws2.cell(row=row_tr, column=3).fill = YELLOW_FILL
ws2.cell(row=row_tr, column=3).number_format = '#,##0'

row_tr2 = row_tr + 1
ws2.merge_cells(f'A{row_tr2}:B{row_tr2}')
ws2.cell(row=row_tr2, column=1, value="期末报酬率").font = BOLD_BLACK
ws2.cell(row=row_tr2, column=3).font = BLUE_FONT; ws2.cell(row=row_tr2, column=3).number_format = '0.00%'

row_tr3 = row_tr2 + 1
ws2.merge_cells(f'A{row_tr3}:C{row_tr3}')
ws2.cell(row=row_tr3, column=1, value="期末转售收益折现").font = BOLD_BLACK
col_c_tr = get_column_letter(3)
formula_tr = f"=ROUND(IF({col_c_tr}{row_tr}*{col_c_tr}{row_tr2}=0,0,{col_c_tr}{row_tr}/(1+{col_c_tr}{row_tr2})^C3),0)"
ws2.cell(row=row_tr3, column=4, value=formula_tr).font = BLACK_FONT
ws2.cell(row=row_tr3, column=4).number_format = '#,##0'

row_total = row_tr3 + 2
ws2.merge_cells(f'A{row_total}:C{row_total}')
ws2.cell(row=row_total, column=1, value="收益总价值 (元/m²)").font = Font(name="Arial", size=12, color="000000", bold=True)
ws2.cell(row=row_total, column=4, value=f"=ROUND(D{row_val}+D{row_tr3},0)").font = Font(name="Arial", size=12, color="000000", bold=True)
ws2.cell(row=row_total, column=4).number_format = '#,##0'; ws2.cell(row=row_total, column=4).fill = YELLOW_FILL

# 数据验证
dv_rate = DataValidation(type="decimal", operator="between", formula1="0", formula2="0.30")
dv_rate.error = "报酬率应在0~30%之间"; ws2.add_data_validation(dv_rate)
dv_rate.add(ws2['C2'])

dv_period = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="1")
dv_period.error = "收益期必须≥1年"; ws2.add_data_validation(dv_period)
dv_period.add(ws2['C3'])

# Sheet 保护
ws1.protection.sheet = True; ws1.protection.password = ""
ws2.protection.sheet = True; ws2.protection.password = ""

out_path = os.path.join(OUTPUT_DIR, "收益法测算_模板.xlsx")
wb.save(out_path)
print(f"✅ 收益法模板已生成: {out_path}")
print(f"   Sheet 1: 净收益测算 (21行)")
print(f"   Sheet 2: 收益价值 (60年折现行 + 期末转售)")
print(f"   数据验证: 空置率0-100%, 报酬率0-30%, 收益期≥1")
