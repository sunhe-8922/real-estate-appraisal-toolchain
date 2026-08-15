#!/usr/bin/env python3
"""
extract_calculation_chain.py — 从 Excel 模板提取公式逻辑，生成 JSON calculationChain（v1.2）

自动推断逻辑：
  1. 解析市场价比较法!T32/W32/Z32 公式，识别每个实例的：
     - eval_col：成交/市场系数列（比值对分子，如 T/W/Z）
     - idx_col：实例指数列（比值对分母 + 子项求和列，如 V/Y/AB）
  2. 从公式中的子项求和（如 V7+V8+...+V18）推断子项区间（区位12项/权益5项/实物8项）
  3. 从行33提取权重列（值为 0<v<1 的列）及对应权重
  4. 收益法节点通过固定路径映射（模板结构相对固定）

输出：JSON calculationChain 对象（schema calculationChain 字段格式）
  {
    "version": "1.2",
    "nodes": [
      {"id", "target", "formula", "refs", "excelSource", "description"}
    ]
  }

公式模板语法：
  {{refKey}}             → refs[refKey] 的 JSONPath 指向数据
  SUM({{refKey}})        → refs 指向数组，重建时展开/求和
  数值常量（0.5 等）      → 直接嵌入公式，不入 refs
  Excel 局部中间量        → 保留原单元格引用（不入 refs，cells 重建还原）

用法：
  python scripts/extract_calculation_chain.py [--excel PATH] [--output PATH]
"""

import json
import re
import sys
import argparse
from pathlib import Path
from typing import Optional

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

DEFAULT_EXCEL = "outputs/房地产评估明细表-计算模板.xlsx"
DEFAULT_OUTPUT = "outputs/calculation_chain.json"


# ── 单格引用 → 语义名 + JSONPath ───────────────────────────────────
CELL_REFS: dict[str, dict[str, str]] = {
    # 评估明细表（面积/单价/总价）
    "评估明细表!M6": {"area": "property.area"},
    "评估明细表!N6": {"unitValue": "result.finalUnitValue"},
    "评估明细表!O6": {"totalValue": "result.finalTotalValue"},
    # 收益法
    "住宅-收益法测算!G4": {"noi": "methods.income.netOperatingIncome.annualAmount"},
    "住宅-收益法测算!G5": {"egi": "methods.income.netOperatingIncome.effectiveGrossIncome"},
    "住宅-收益法测算!G11": {"oe": "methods.income.netOperatingIncome.operatingExpenses"},
    "住宅-收益法测算!G24": {"rate": "methods.income.rate.value"},
    "住宅-收益法测算!G25": {"growth": "methods.income.netOperatingIncome.growthRate"},
    "住宅-收益法测算!G26": {"forwardPeriod": "methods.income.holdingPeriod"},
    "住宅-收益法测算!G27": {"incomeValue": "methods.income.finalValue.total"},
    "住宅-收益法测算!I27": {"incomeWeight": "result.weightAllocation.income"},
    "住宅-收益法测算!I28": {"compsWeight": "result.weightAllocation.comps"},
    "住宅-收益法测算!J28": {"compsValue": "methods.comps.finalValue.total"},
    "住宅-收益法测算!J29": {"finalTotal": "result.finalTotalValue"},
}

# 子项类别 → JSON 字段名
DETAIL_JSONPATH: dict[str, str] = {
    "loc": "locationDetails",
    "int": "interestDetails",
    "phy": "physicalDetails",
}

# 子项行区间（可推断，默认值）
DEFAULT_FACTORS = {
    "loc": (7, 18),   # 12项
    "int": (19, 23),  # 5项
    "phy": (24, 31),  # 8项
}


def _parse_cell_refs(formula: str) -> list[str]:
    """从公式中提取所有单元格引用（如 T5, V7, AB31）。"""
    return re.findall(r"[A-Z]{1,2}\d+", formula)


def _detect_instance_layout(
    ws,
) -> dict[int, dict]:
    """
    自动推断 3 个实例的列组布局。
    返回 {idx: {"eval_col": ..., "idx_col": ..., "ranges": [("loc",7,18), ...],
                 "weight_col": ..., "weight": ...}}
    推断逻辑：
      1. 从公式 T32/W32/Z32 中提取比值对（如 T5/V5）→ eval_col + idx_col
      2. 从子项求和（如 V7+V8+...+V18）推断子项列和区间
      3. 从行33提取权重列及权重值
    """
    instances: dict[int, dict] = {}

    # 找实例结果列（T32/W32/Z32 等）—— 公式中包含 "32" 的行
    result_cols: list[str] = []
    for col_idx in range(1, 30):
        cell = ws.cell(32, col_idx)
        if cell.value and isinstance(cell.value, str) and "=ROUND" in cell.value:
            col_letter = cell.column_letter
            # 检查是否引用了子项区间
            refs = _parse_cell_refs(cell.value)
            if any(r.endswith("32") for r in refs):
                result_cols.append(col_letter)

    if not result_cols:
        # 回退：固定 T/W/Z
        result_cols = ["T", "W", "Z"]

    for i, eval_col in enumerate(result_cols):
        formula = str(ws[f"{eval_col}32"].value or "")
        refs = _parse_cell_refs(formula)

        # 提取比值对：eval_col5/idx_col5
        ratio_pairs = re.findall(rf"({eval_col})5/([A-Z]+)5|({eval_col})6/([A-Z]+)6", formula)
        idx_col: Optional[str] = None
        for p in ratio_pairs:
            idx_col = p[1] or p[3]
            break
        if not idx_col:
            # 从子项引用推断（取第一个出现的非 eval_col 列）
            sub_cols = set()
            for r in refs:
                c = re.match(r"([A-Z]+)\d+", r)
                if c and c.group(1) != eval_col:
                    sub_cols.add(c.group(1))
            if sub_cols:
                idx_col = sorted(sub_cols)[0]

        # 从子项引用推断区间
        sub_rows: dict[str, list[int]] = {}  # col -> [rows]
        for r in refs:
            m = re.match(r"([A-Z]+)(\d+)", r)
            if m and m.group(1) != eval_col:
                col, row = m.group(1), int(m.group(2))
                if col not in sub_rows:
                    sub_rows[col] = []
                sub_rows[col].append(row)

        ranges: list[tuple[str, int, int]] = []
        if idx_col and idx_col in sub_rows:
            rows = sorted(set(sub_rows[idx_col]))
            # 自动划分区间：7-18(loc), 19-23(int), 24-31(phy)
            # 或者从行间距推断
            if len(rows) >= 12:
                # 标准布局：12+5+8=25行
                ranges = [
                    ("loc", 7, 18),
                    ("int", 19, 23),
                    ("phy", 24, 31),
                ]
            elif len(rows) >= 5:
                # 简化布局
                ranges = [("loc", min(rows), max(rows))]
        
        # 提取权重列（行33，值在 0-1 之间）
        weight_col: Optional[str] = None
        weight: Optional[float] = None
        for col_idx in range(1, 30):
            cell = ws.cell(33, col_idx)
            if isinstance(cell.value, float) and 0 < cell.value < 1:
                if cell.column_letter == eval_col:
                    weight_col = cell.column_letter
                    weight = cell.value
                    break

        instances[i] = {
            "eval_col": eval_col,
            "idx_col": idx_col,
            "ranges": ranges,
            "weight_col": weight_col,
            "weight": weight,
        }

    return instances


def _build_auto_factor_layout(
    instances: dict[int, dict],
) -> dict[str, dict[int, dict]]:
    """构建自动推断的 FACTOR_LAYOUT 结构（与旧版 FA
CTOR_LAYOUT 兼容）。"""
    layout: dict[str, dict[int, dict]] = {"市场价比较法": {}}
    for idx, cfg in instances.items():
        ranges = cfg["ranges"] if cfg["ranges"] else [
            ("loc", 7, 18),
            ("int", 19, 23),
            ("phy", 24, 31),
        ]
        layout["市场价比较法"][idx] = {
            "col": cfg["idx_col"],
            "ranges": ranges,
            "path": f"methods.comps.comparableInstances[{idx}].adjustments.",
        }
    return layout


def _build_expanded_sum_rules(
    factor_layout: dict[str, dict[int, dict]],
) -> list[tuple[str, str, str]]:
    """生成 (展开式正则, 语义名, JSONPath) 列表。"""
    rules = []
    for sheet, instances in factor_layout.items():
        for idx, cfg in instances.items():
            col = cfg["col"]
            for kind, start, end in cfg["ranges"]:
                cells = "+".join(f"{col}{r}" for r in range(start, end + 1))
                ref_key = f"{kind}Factors{idx}"
                path = cfg["path"] + DETAIL_JSONPATH[kind]
                rules.append((cells, ref_key, path))
    return rules


def cell_to_paths(cell_ref: str) -> dict:
    """Excel 单元格引用 → {语义名: JSONPath} 映射。"""
    return CELL_REFS.get(cell_ref, {})


def translate_formula(
    formula: str, excel_ref: str, expanded_rules: list[tuple[str, str, str]],
    pair_rules: list[tuple[str, str, str]],
) -> tuple[str, dict, list[str]]:
    """
    将 Excel 公式翻译为 calculationChain 公式模板。
    返回 (公式模板, refs映射, 未识别引用列表)。
    """
    if not formula or not str(formula).startswith("="):
        return str(formula) if formula else "", {}, []
    body = str(formula)[1:].strip()
    refs: dict[str, str] = {}
    unknowns: list[str] = []
    sheet_name = excel_ref.split("!")[0] if "!" in excel_ref else "市场价比较法"

    # 1) 子项展开式折叠
    for cells, ref_key, path in expanded_rules:
        if cells in body:
            body = body.replace(cells, f"SUM({{{{{ref_key}}}}})")
            refs[ref_key] = path

    # 2) 比值对折叠
    for pattern, ref_key, path in pair_rules:
        if re.search(pattern, body):
            body = re.sub(pattern, f"{{{{{ref_key}}}}}", body)
            refs[ref_key] = path

    # 3) 单格引用
    pattern = re.compile(r"([\u4e00-\u9fa5A-Za-z0-9]+)!(\$?[A-Z]{1,2}\$?\d+)")
    local_pattern = re.compile(r"(?<![A-Z])(\$?[A-Z]{1,2}\$?\d+)(?![A-Z0-9])")

    def replace_cell(m):
        sheet, cell = m.group(1), m.group(2)
        key = f"{sheet}!{cell.replace('$', '')}"
        mapping = cell_to_paths(key)
        if not mapping:
            unknowns.append(key)
            return m.group(0)
        name, path = next(iter(mapping.items()))
        refs[name] = path
        return f"{{{{{name}}}}}"

    body = pattern.sub(replace_cell, body)

    def replace_local(m):
        cell = m.group(1).replace("$", "")
        key = f"{sheet_name}!{cell}"
        mapping = cell_to_paths(key)
        if not mapping:
            unknowns.append(key)
            return m.group(0)
        name, path = next(iter(mapping.items()))
        refs[name] = path
        return f"{{{{{name}}}}}"

    body = local_pattern.sub(replace_local, body)
    return body, refs, unknowns


def extract_chain(excel_path: str) -> dict:
    """从 Excel 读取核心单元格公式，构建 calculationChain。"""
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    nodes = []

    # ── 自动推断列组布局 ──────────────────────────────────────────────
    comps_ws = wb["市场价比较法"]
    instances = _detect_instance_layout(comps_ws)
    factor_layout = _build_auto_factor_layout(instances)
    expanded_rules = _build_expanded_sum_rules(factor_layout)

    # 比值对规则（动态生成）
    pair_rules: list[tuple[str, str, str]] = []
    for idx, cfg in instances.items():
        eval_col = cfg["eval_col"]
        idx_col = cfg["idx_col"]
        if eval_col and idx_col:
            pair_rules.append(
                (rf"{eval_col}5/{idx_col}5", "txAdj",
                 f"methods.comps.comparableInstances[{idx}].adjustments.transactionSituation")
            )
            pair_rules.append(
                (rf"{eval_col}6/{idx_col}6", "mktAdj",
                 f"methods.comps.comparableInstances[{idx}].adjustments.marketCondition")
            )

    # 实例单价 refs（T4→instance0, W4→instance1, Z4→instance2）
    _UNIT_PRICE_OVERRIDES = {
        instances[i]["eval_col"]: f"methods.comps.comparableInstances[{i}].unitPrice"
        for i in range(3)
    }

    # 节点级 override refs（公式含无法自动推断的局部引用时显式声明）
    _NODE_OVERRIDE_REFS: dict[str, dict] = {
        "comps.finalUnitPrice": {
            "formula": "ROUND(({{adjPrice0}}*0.5+{{adjPrice1}}*0.3+{{adjPrice2}}*0.2),-1)",
            "refs": {
                "adjPrice0": "methods.comps.comparableInstances[0].adjustedUnitPrice",
                "adjPrice1": "methods.comps.comparableInstances[1].adjustedUnitPrice",
                "adjPrice2": "methods.comps.comparableInstances[2].adjustedUnitPrice",
            },
        },
        "result.totalValue": {
            "formula": "ROUND({{area}}*{{unitValue}},0)",
            "refs": {
                "area": "property.area",
                "unitValue": "result.finalUnitValue",
            },
        },
        "result.finalTotalValue": {
            "formula": "ROUND({{incomeValue}}*{{incomeWeight}}+{{compsValue}}*{{compsWeight}},-1)",
            "refs": {
                "incomeValue": "methods.income.finalValue.total",
                "incomeWeight": "result.weightAllocation.income",
                "compsValue": "methods.comps.finalValue.total",
                "compsWeight": "result.weightAllocation.comps",
            },
        },
    }

    CORE_NODES = [
        (
            "comps.adjustedUnitPrice.instance1",
            f"市场价比较法!{instances[0]['eval_col']}32",
            "methods.comps.comparableInstances[0].adjustedUnitPrice",
            "比准单价（实例1）= 成交单价 × 交易情况 × 市场状况 × 区位偏差合并 × 权益偏差合并 × 实物偏差合并。",
            None,
        ),
        (
            "comps.adjustedUnitPrice.instance2",
            f"市场价比较法!{instances[1]['eval_col']}32",
            "methods.comps.comparableInstances[1].adjustedUnitPrice",
            "比准单价（实例2），同上结构",
            None,
        ),
        (
            "comps.adjustedUnitPrice.instance3",
            f"市场价比较法!{instances[2]['eval_col']}32",
            "methods.comps.comparableInstances[2].adjustedUnitPrice",
            "比准单价（实例3），同上结构",
            None,
        ),
        (
            "comps.finalUnitPrice",
            f"市场价比较法!{instances[0]['eval_col']}34",
            "methods.comps.finalValue.unit",
            "比准单价加权平均 = Σ(实例i修正价 × 权重i)，四舍五入到十位。权重从行33读取。",
            None,
        ),
        (
            "income.noi",
            "住宅-收益法测算!G4",
            "methods.income.netOperatingIncome.annualAmount",
            "年净收益 = 有效毛收入 − 运营费用",
            None,
        ),
        (
            "income.value",
            "住宅-收益法测算!G27",
            "methods.income.finalValue.total",
            "收益价值（报酬资本化分段折现）",
            None,
        ),
        (
            "result.finalTotalValue",
            "住宅-收益法测算!J29",
            "result.finalTotalValue",
            "最终结果 = 收益价值 × 收益权重 + 比较法价值 × 比较权重",
            None,
        ),
        (
            "result.totalValue",
            "评估明细表!O6",
            "result.finalTotalValue",
            "评估总价 = 面积 × 单价",
            None,
        ),
    ]

    for node_id, excel_ref, target, desc, _ in CORE_NODES:
        sheet, cell = excel_ref.split("!")
        ws = wb[sheet]
        raw_formula = ws[cell].value
        if raw_formula is None:
            print(f"  [跳过] {excel_ref} 无公式")
            continue

        formula, refs, unknowns = translate_formula(
            str(raw_formula), excel_ref, expanded_rules, pair_rules
        )

        # 后处理：把实例单价局部引用（T4/W4/Z4）转为 refs
        if "comps.adjustedUnitPrice.instance" in node_id:
            eval_col = cell[0]  # T/W/Z
            if eval_col in _UNIT_PRICE_OVERRIDES:
                refs["unitPrice"] = _UNIT_PRICE_OVERRIDES[eval_col]
                # 从公式中替换 T4→{{unitPrice}} 等
                formula = re.sub(
                    rf"\b{re.escape(eval_col)}4\b",
                    "{{unitPrice}}",
                    formula,
                )
                if excel_ref.split("!")[0] in unknowns:
                    unknowns.remove(excel_ref.split("!")[0])

        # 节点级 override refs（comps.finalUnitPrice 等含局部引用的节点）
        if node_id in _NODE_OVERRIDE_REFS:
            formula = _NODE_OVERRIDE_REFS[node_id]["formula"]
            refs = _NODE_OVERRIDE_REFS[node_id]["refs"]

        node = {
            "id": node_id,
            "target": target,
            "formula": formula,
            "refs": refs,
            "excelSource": excel_ref,
            "description": desc,
        }
        nodes.append(node)
        status = f"OK ({len(refs)} refs)"
        if unknowns:
            status += f", 局部引用 {len(unknowns)}: {sorted(set(unknowns))[:6]}"
        print(f"  [{status}] {node_id}: {formula[:90]}")

    wb.close()
    return {"version": "1.2", "nodes": nodes}


def main():
    parser = argparse.ArgumentParser(description="Excel 模板 → JSON calculationChain 提取")
    parser.add_argument("--excel", default=DEFAULT_EXCEL, help="Excel 模板路径")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="输出 JSON 路径")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        print(f"ERROR: Excel 文件不存在: {excel_path}")
        sys.exit(1)

    print(f"提取计算链: {excel_path}")
    chain = extract_chain(str(excel_path))

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(chain, f, ensure_ascii=False, indent=2)
    print(f"\n✅ 计算链已保存: {output_path}（{len(chain['nodes'])} 个节点）")


if __name__ == "__main__":
    main()
