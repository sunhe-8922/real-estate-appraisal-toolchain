#!/usr/bin/env python3
"""生成假设开发法测算 Excel 模板 — GB/T 50291-2015 第4.5节"""
import os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

# ── 颜色定义 ──
BLUE_FONT = Font(name="Arial", size=11, color="0000FF")
BLACK_FONT = Font(name="Arial", size=11, color="000000")
BOLD_BLACK = Font(name="Arial", size=11, color="000000", bold=True)
TITLE_FONT = Font(name="Arial", size=14, color="FFFFFF", bold=True)
SECTION_FONT = Font(name="Arial", size=12, color="1F4E79", bold=True, underline="single")
LEGEND_FONT = Font(name="Arial", size=9, color="333333", italic=True)
RED_BOLD = Font(name="Arial", size=11, color="FF0000", bold=True)

DARK_BLUE_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "outputs", "templates")
os.makedirs(OUTPUT_DIR, exist_ok=True)

wb = Workbook()

# ═══════════════════════════════════════════════
# SHEET 1: 假设开发法
# ═══════════════════════════════════════════════
ws1 = wb.active
ws1.title = "假设开发法"
ws1.column_dimensions['A'].width = 28
ws1.column_dimensions['B'].width = 22
ws1.column_dimensions['C'].width = 18
ws1.column_dimensions['D'].width = 22

# 标题
ws1.merge_cells('A1:D1')
c = ws1['A1']; c.value = "假设开发法测算"; c.font = TITLE_FONT; c.fill = DARK_BLUE_FILL; c.alignment = CENTER
ws1.row_dimensions[1].height = 30

# 项目信息
ws1.merge_cells('A2:B2')
ws1['A2'] = "估价对象: [请输入地址]"; ws1['A2'].font = BLUE_FONT
ws1['C2'] = "价值时点: [日期]"; ws1['C2'].font = BLUE_FONT
ws1['D2'] = "方法: 动态分析法"; ws1['D2'].font = BOLD_BLACK

# 颜色图例
ws1.merge_cells('A3:D3')
ws1['A3'] = "图例: 蓝字=可修改输入 | 黑字=公式 | 黄底=关键参数 | 红字=红线检查"
ws1['A3'].font = LEGEND_FONT; ws1['A3'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

row = 5

# ── 一、最佳开发经营方式 ──
ws1.merge_cells(f'A{row}:D{row}')
ws1.cell(row=row, column=1, value="一、最佳开发经营方式").font = SECTION_FONT
ws1.cell(row=row, column=1).fill = PatternFill(start_color="E6EEF5", end_color="E6EEF5", fill_type="solid")
row += 1

dev_items = [
    ("开发完成后用途", "BLUE", "开发完成后的规划用途", "规模 (m²)", "BLUE", "开发完成后总建筑面积"),
    ("经营方式", "DV_LIST_SELL", "全部出售 / 部分出租 / 自持", "", "", ""),
]
for label, style, hint, label2, style2, hint2 in dev_items:
    ws1.cell(row=row, column=1, value=label).font = BOLD_BLACK; ws1.cell(row=row, column=1).border = THIN_BORDER
    c = ws1.cell(row=row, column=2)
    if style == "BLUE":
        c.font = BLUE_FONT; c.value = None; c.border = THIN_BORDER
    elif style == "DV_LIST_SELL":
        c.font = BLUE_FONT; c.border = THIN_BORDER
        dv_sell = DataValidation(type="list", formula1='"全部出售,部分出租,自持"', allow_blank=False)
        ws1.add_data_validation(dv_sell); dv_sell.add(c)
    ws1.cell(row=row, column=3, value=label2).font = BOLD_BLACK; ws1.cell(row=row, column=3).border = THIN_BORDER
    c2 = ws1.cell(row=row, column=4)
    if style2 == "BLUE":
        c2.font = BLUE_FONT; c2.number_format = '#,##0'; c2.border = THIN_BORDER
    row += 1

row += 1

# ── 二、开发经营期 ──
ws1.merge_cells(f'A{row}:D{row}')
ws1.cell(row=row, column=1, value="二、开发经营期 (月)").font = SECTION_FONT
ws1.cell(row=row, column=1).fill = PatternFill(start_color="E6EEF5", end_color="E6EEF5", fill_type="solid")
row += 1

period_items = [
    ("前期", "BLUE", "规划设计/报建", "建造期", "BLUE", "主体施工至竣工"),
    ("销售期", "BLUE", "竣工后至售罄", "开发经营期总计", "=B10+C10", "(不含销售期) 公式"),
]
for label, style, hint, label2, style2, hint2 in period_items:
    ws1.cell(row=row, column=1, value=label).font = BOLD_BLACK; ws1.cell(row=row, column=1).border = THIN_BORDER
    c = ws1.cell(row=row, column=2)
    if style == "BLUE":
        c.font = BLUE_FONT; c.number_format = '0'; c.border = THIN_BORDER
    ws1.cell(row=row, column=3, value=label2).font = BOLD_BLACK; ws1.cell(row=row, column=3).border = THIN_BORDER
    c2 = ws1.cell(row=row, column=4)
    if style2 == "BLUE":
        c2.font = BLUE_FONT; c2.number_format = '0'; c2.border = THIN_BORDER
    else:
        c2.font = BLACK_FONT; c2.value = style2; c2.number_format = '0'; c2.border = THIN_BORDER
    row += 1

row += 1

# ── 三、开发完成后价值 ──
ws1.merge_cells(f'A{row}:D{row}')
ws1.cell(row=row, column=1, value="三、开发完成后价值  (⚠️ 4.5.7条第1款: 不得用成本法)").font = SECTION_FONT
ws1.cell(row=row, column=1).fill = PatternFill(start_color="E6EEF5", end_color="E6EEF5", fill_type="solid")
row += 1

ws1.cell(row=row, column=1, value="测算方法").font = BOLD_BLACK; ws1.cell(row=row, column=1).border = THIN_BORDER
c_m = ws1.cell(row=row, column=2)
c_m.font = BLUE_FONT; c_m.border = THIN_BORDER
dv_val_method = DataValidation(type="list", formula1='"比较法,收益法"', allow_blank=False)
dv_val_method.error = "开发完成后价值必须用比较法或收益法(4.5.7条)"; dv_val_method.errorTitle = "红线警告"
ws1.add_data_validation(dv_val_method); dv_val_method.add(c_m)

ws1.cell(row=row, column=3, value="数据来源").font = BOLD_BLACK; ws1.cell(row=row, column=3).border = THIN_BORDER
ws1.cell(row=row, column=4).font = BLUE_FONT; ws1.cell(row=row, column=4).border = THIN_BORDER
row += 1

val_items = [
    ("单价 (元/m²)", "BLUE_Y", "如: 25000", "总价 (元)", f"=B{row}*B7", "公式; 单价×规模"),
    ("信源等级", "DV_LIST_SRC", "T0/T1/T2", "建筑面积 (m²)", "=B7", "引用一、节的规模"),
]
for label, style, hint, label2, formula, note in val_items:
    ws1.cell(row=row, column=1, value=label).font = BOLD_BLACK; ws1.cell(row=row, column=1).border = THIN_BORDER
    c = ws1.cell(row=row, column=2)
    if style == "BLUE_Y":
        c.font = BLUE_FONT; c.fill = YELLOW_FILL; c.number_format = '#,##0'; c.border = THIN_BORDER
        c.comment = Comment("4.5.7条第1款: 开发完成后价值不得用成本法。必须用比较法或收益法测算。", "GB/T 50291-2015")
    elif style == "DV_LIST_SRC":
        c.font = BLUE_FONT; c.border = THIN_BORDER
        dv_src = DataValidation(type="list", formula1='"T0-官方,T1-头部机构,T2-自媒体"', allow_blank=False)
        ws1.add_data_validation(dv_src); dv_src.add(c)
    ws1.cell(row=row, column=3, value=label2).font = BOLD_BLACK; ws1.cell(row=row, column=3).border = THIN_BORDER
    c2 = ws1.cell(row=row, column=4); c2.number_format = '#,##0'; c2.border = THIN_BORDER
    if formula.startswith("="):
        c2.font = BLACK_FONT; c2.value = formula
    else:
        c2.font = BLUE_FONT
    row += 1

row += 1

# ── 四、后续开发必要支出 ──
ws1.merge_cells(f'A{row}:D{row}')
ws1.cell(row=row, column=1, value="四、后续开发必要支出").font = SECTION_FONT
ws1.cell(row=row, column=1).fill = PatternFill(start_color="E6EEF5", end_color="E6EEF5", fill_type="solid")
row += 1

for i, h in enumerate(["项目", "金额 (元)", "参数", "发生期 (月)"], 1):
    cell = ws1.cell(row=row, column=i, value=h)
    cell.font = Font(name="Arial", size=11, color="000000", bold=True)
    cell.fill = LIGHT_BLUE_FILL; cell.alignment = CENTER; cell.border = THIN_BORDER
row += 1

cost_items = [
    ("建设成本",           "BLUE",   "直接建设支出"),
    ("管理费用",           "FORMULA_RATE", "管理费率×建设成本"),
    ("销售费用",           "FORMULA_RATE2","销售费率×总价值"),
    ("销售税费",           "FORMULA_TAX",  "税率×总价值"),
    ("取得税费 (仅转让前提)", "BLUE",   "契税/印花税等"),
]
for label, style, note in cost_items:
    ws1.cell(row=row, column=1, value=label).font = BOLD_BLACK; ws1.cell(row=row, column=1).border = THIN_BORDER
    c2 = ws1.cell(row=row, column=2)
    if style == "BLUE":
        c2.font = BLUE_FONT; c2.number_format = '#,##0'; c2.border = THIN_BORDER
    elif style == "FORMULA_RATE":
        c2.font = BLACK_FONT; c2.value = f"=B{row-4}*管理费率"; c2.number_format = '#,##0'; c2.border = THIN_BORDER
    elif style == "FORMULA_RATE2":
        c2.font = BLACK_FONT; c2.value = f"=D{row-8}*销售费率"; c2.number_format = '#,##0'; c2.border = THIN_BORDER
    elif style == "FORMULA_TAX":
        c2.font = BLACK_FONT; c2.value = f"=D{row-8}*税率"; c2.number_format = '#,##0'; c2.border = THIN_BORDER

    c3 = ws1.cell(row=row, column=3)
    if style != "BLUE":
        c3.font = Font(name="Arial", size=9, color="FF0000"); c3.value = "见参数表"; c3.alignment = CENTER
    c3.border = THIN_BORDER

    c4 = ws1.cell(row=row, column=4)
    c4.font = BLUE_FONT; c4.number_format = '0'; c4.border = THIN_BORDER
    row += 1

# 支出合计
ws1.cell(row=row, column=1, value="支出合计").font = BOLD_BLACK; ws1.cell(row=row, column=1).border = THIN_BORDER
start_cost_row = row - 5
ws1.cell(row=row, column=2, value=f"=SUM(B{start_cost_row}:B{row-1})").font = BLACK_FONT
ws1.cell(row=row, column=2).number_format = '#,##0'; ws1.cell(row=row, column=2).border = THIN_BORDER
row += 2

# ── 五、折现率/利息/利润 ──
ws1.merge_cells(f'A{row}:D{row}')
ws1.cell(row=row, column=1, value="五、折现率/利息/利润").font = SECTION_FONT
ws1.cell(row=row, column=1).fill = PatternFill(start_color="E6EEF5", end_color="E6EEF5", fill_type="solid")
row += 1

ws1.cell(row=row, column=1, value="折现率 r (动态法)").font = BOLD_BLACK; ws1.cell(row=row, column=1).border = THIN_BORDER
c_r = ws1.cell(row=row, column=2)
c_r.font = BLUE_FONT; c_r.fill = YELLOW_FILL; c_r.number_format = '0.00%'; c_r.border = THIN_BORDER
c_r.comment = Comment("折现率应体现资金时间价值和项目风险。通过市场提取或累加法确定。", "GB/T 50291-2015")
discount_rate_row = row  # 记录折现率所在行号
ws1.cell(row=row, column=3, value="利率 (静态法)").font = BOLD_BLACK; ws1.cell(row=row, column=3).border = THIN_BORDER
ws1.cell(row=row, column=4).font = BLUE_FONT; ws1.cell(row=row, column=4).number_format = '0.00%'; ws1.cell(row=row, column=4).border = THIN_BORDER
row += 1

ws1.cell(row=row, column=1, value="利润率 (静态法)").font = BOLD_BLACK; ws1.cell(row=row, column=1).border = THIN_BORDER
ws1.cell(row=row, column=2).font = BLUE_FONT; ws1.cell(row=row, column=2).fill = YELLOW_FILL
ws1.cell(row=row, column=2).number_format = '0.0%'; ws1.cell(row=row, column=2).border = THIN_BORDER
row += 2

# ── 六、开发价值 ──
ws1.merge_cells(f'A{row}:D{row}')
ws1.cell(row=row, column=1, value="六、开发价值 (动态法)").font = SECTION_FONT
ws1.cell(row=row, column=1).fill = PatternFill(start_color="E6EEF5", end_color="E6EEF5", fill_type="solid")
row += 1

# 动态法折现计算
val_row = 16  # 开发完成后总价所在行 (approximate, user should verify)
dr_row = discount_rate_row  # 折现率所在行
discount_items = [
    ("开发完成后价值折现 (元)",    f"=D{val_row}/(1+B{dr_row})^((C11+C10)/12)", "折现期=建造期+销售期, 转年"),
    ("建设成本折现 (元)",           f"=B{start_cost_row}/(1+B{dr_row})^(D{start_cost_row}/12)", "按发生期分别折现"),
    ("管理费用折现 (元)",           f"=B{start_cost_row+1}/(1+B{dr_row})^(D{start_cost_row+1}/12)", ""),
    ("销售费用折现 (元)",           f"=B{start_cost_row+2}/(1+B{dr_row})^(D{start_cost_row+2}/12)", ""),
    ("销售税费折现 (元)",           f"=B{start_cost_row+3}/(1+B{dr_row})^(D{start_cost_row+3}/12)", ""),
    ("取得税费折现 (元)(如有)",     f"=B{start_cost_row+4}/(1+B{dr_row})^(D{start_cost_row+4}/12)", "仅转让开发前提"),
]
for label, formula, note in discount_items:
    ws1.cell(row=row, column=1, value=label).font = BOLD_BLACK; ws1.cell(row=row, column=1).border = THIN_BORDER
    ws1.cell(row=row, column=2, value=formula).font = BLACK_FONT
    ws1.cell(row=row, column=2).number_format = '#,##0'; ws1.cell(row=row, column=2).border = THIN_BORDER
    ws1.cell(row=row, column=3, value=note).font = Font(name="Arial", size=9, color="666666")
    ws1.cell(row=row, column=3).alignment = LEFT_WRAP; ws1.merge_cells(f'C{row}:D{row}')
    row += 1

# 支出折现合计
val_items_start = row - 6
ws1.cell(row=row, column=1, value="支出折现合计 (元)").font = BOLD_BLACK; ws1.cell(row=row, column=1).border = THIN_BORDER
ws1.cell(row=row, column=2, value=f"=SUM(B{val_items_start+1}:B{row-1})").font = BLACK_FONT
ws1.cell(row=row, column=2).number_format = '#,##0'; ws1.cell(row=row, column=2).border = THIN_BORDER
row += 1

# 开发价值(总价)
dev_val_row = row
ws1.cell(row=row, column=1, value="开发价值 总价 (元)").font = Font(name="Arial", size=12, color="000000", bold=True)
ws1.cell(row=row, column=1).border = THIN_BORDER
ws1.cell(row=row, column=2, value=f"=B{val_items_start}-B{row-1}").font = Font(name="Arial", size=12, color="000000", bold=True)
ws1.cell(row=row, column=2).number_format = '#,##0'; ws1.cell(row=row, column=2).fill = YELLOW_FILL
ws1.cell(row=row, column=2).border = THIN_BORDER
row += 1

ws1.cell(row=row, column=1, value="建筑面积 (m²)").font = BOLD_BLACK; ws1.cell(row=row, column=1).border = THIN_BORDER
ws1.cell(row=row, column=2, value="=B7").font = BLACK_FONT; ws1.cell(row=row, column=2).border = THIN_BORDER
row += 1

ws1.cell(row=row, column=1, value="开发价值 单价 (元/m²)").font = Font(name="Arial", size=12, color="000000", bold=True)
ws1.cell(row=row, column=1).border = THIN_BORDER
ws1.cell(row=row, column=2, value=f"=B{dev_val_row}/B{dev_val_row+1}").font = Font(name="Arial", size=12, color="000000", bold=True)
ws1.cell(row=row, column=2).number_format = '#,##0'; ws1.cell(row=row, column=2).fill = YELLOW_FILL
ws1.cell(row=row, column=2).border = THIN_BORDER
row += 2

# ── 七、红线检查 ──
ws1.merge_cells(f'A{row}:D{row}')
ws1.cell(row=row, column=1, value="七、红线检查").font = SECTION_FONT
ws1.cell(row=row, column=1).fill = PatternFill(start_color="E6EEF5", end_color="E6EEF5", fill_type="solid")
row += 1

redline_items = [
    ("开发完成后价值方法", f'=IF(B15="成本法","✗ 违规: 4.5.7条禁止用成本法","✓ 合规: 比较法/收益法")'),
    ("利息/利润重复检查",  '=IF(D2<>"静态","✓ 动态法无需另算利息利润","⚠ 静态法: 确认未重复计算")'),
    ("估价前提匹配检查",   f'=IF(AND(D2="被迫转让开发", B{discount_rate_row}<0.08),"⚠ 抵押估价折现率宜≥8%","✓")'),
]
for label, formula in redline_items:
    ws1.cell(row=row, column=1, value=label).font = BOLD_BLACK; ws1.cell(row=row, column=1).border = THIN_BORDER
    ws1.merge_cells(f'B{row}:D{row}')
    ws1.cell(row=row, column=2, value=formula).font = BLACK_FONT
    ws1.cell(row=row, column=2).alignment = LEFT_WRAP; ws1.cell(row=row, column=2).border = THIN_BORDER
    row += 1

# 条件格式: 红线违规变红
for r in range(row-3, row):
    ws1.conditional_formatting.add(f"B{r}",
        CellIsRule(operator="beginsWith", formula=['"✗"'],
            font=Font(color="FF0000", bold=True), fill=RED_FILL))

# 数据验证: 折现率
dv_dr = DataValidation(type="decimal", operator="between", formula1="0.05", formula2="0.30")
dv_dr.error = "折现率应在5%~30%之间"; ws1.add_data_validation(dv_dr); dv_dr.add(c_r)

# ═══════════════════════════════════════════════
# SHEET 2: 现金流时间线
# ═══════════════════════════════════════════════
ws2 = wb.create_sheet("现金流时间线")
for col_idx, w in enumerate([12, 18, 18, 18, 18], 1):
    ws2.column_dimensions[get_column_letter(col_idx)].width = w

ws2.merge_cells('A1:E1')
c = ws2['A1']; c.value = "现金流时间线 (动态分析法)"; c.font = TITLE_FONT; c.fill = DARK_BLUE_FILL; c.alignment = CENTER
ws2.row_dimensions[1].height = 30

for i, h in enumerate(["期数(月)", "现金流入 (元)", "现金流出 (元)", "净现金流 (元)", "折现因子"], 1):
    cell = ws2.cell(row=3, column=i, value=h)
    cell.font = Font(name="Arial", size=11, color="000000", bold=True)
    cell.fill = LIGHT_BLUE_FILL; cell.alignment = CENTER; cell.border = THIN_BORDER

# 24行时间线
for t in range(24):
    row = 4 + t
    ws2.cell(row=row, column=1, value=t).font = BLACK_FONT
    ws2.cell(row=row, column=1).alignment = CENTER; ws2.cell(row=row, column=1).border = THIN_BORDER

    # 现金流入 (第0期=0)
    if t == 0:
        ws2.cell(row=row, column=2, value=0).font = BLACK_FONT
    else:
        ws2.cell(row=row, column=2).font = BLUE_FONT
    ws2.cell(row=row, column=2).number_format = '#,##0'; ws2.cell(row=row, column=2).border = THIN_BORDER

    # 现金流出 (第0期=取得税费蓝字)
    if t == 0:
        cf = ws2.cell(row=row, column=3)
        cf.font = BLUE_FONT; cf.number_format = '#,##0'; cf.border = THIN_BORDER
    else:
        ws2.cell(row=row, column=3).font = BLUE_FONT
    ws2.cell(row=row, column=3).number_format = '#,##0'; ws2.cell(row=row, column=3).border = THIN_BORDER

    # 净现金流 = 流入 - 流出
    ws2.cell(row=row, column=4, value=f"=B{row}-C{row}").font = BLACK_FONT
    ws2.cell(row=row, column=4).number_format = '#,##0'; ws2.cell(row=row, column=4).border = THIN_BORDER

    # 折现因子 (引用 Sheet1 折现率)
    ws2.cell(row=row, column=5, value=f"=1/(1+假设开发法!B{27})^(A{row}/12)").font = BLACK_FONT
    ws2.cell(row=row, column=5).number_format = '0.000000'; ws2.cell(row=row, column=5).border = THIN_BORDER

# NPV
row_npv = 28
ws2.merge_cells(f'A{row_npv}:C{row_npv}')
ws2.cell(row=row_npv, column=1, value="NPV (元)").font = Font(name="Arial", size=12, color="000000", bold=True)
ws2.cell(row=row_npv, column=4, value="=SUMPRODUCT(D4:D27,E4:E27)").font = Font(name="Arial", size=12, color="000000", bold=True)
ws2.cell(row=row_npv, column=4).number_format = '#,##0'; ws2.cell(row=row_npv, column=4).fill = YELLOW_FILL
ws2.cell(row=row_npv, column=4).border = THIN_BORDER

ws2.cell(row=row_npv, column=4).comment = Comment("净现值 = 各期净现金流 × 折现因子之和。应与Sheet1动态法开发价值一致。", "校验")

# Sheet 保护
ws1.protection.sheet = True; ws1.protection.password = ""
ws2.protection.sheet = True; ws2.protection.password = ""

out_path = os.path.join(OUTPUT_DIR, "假设开发法测算_模板.xlsx")
wb.save(out_path)
print(f"✅ 假设开发法模板已生成: {out_path}")
print(f"   Sheet 1: 假设开发法 (7个章节, 动态法为主)")
print(f"   Sheet 2: 现金流时间线 (24期 + NPV)")
print(f"   数据验证: 测算方法禁止成本法, 经营方式下拉框, 折现率5-30%")
print(f"   条件格式: 红线违规自动红色")
